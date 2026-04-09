# e:/projects/python/django/teguh/babycare/core/views.py
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, View, FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, update_session_auth_hash
from django.contrib.auth.forms import AuthenticationForm
from django.urls import reverse_lazy
from django.shortcuts import redirect, render, get_object_or_404
from django.db.models import Sum, Q, Count
from django.db import OperationalError
from django.utils import timezone
from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.cache import never_cache
from django.conf import settings
from datetime import datetime, timedelta
from functools import wraps
from urllib.parse import quote

from .models import (
    Pasien, Registrasi, RegistrasiDetail, Pemasukan, Pengeluaran, Notifikasi, 
    Terapis, JenisTerapi, Cabang, User, ProgressTracking, Role, UserRole, 
    TemplatePesan, NOTIFICATION_TYPE_CHOICES, KategoriBarang, BarangInventory, 
    StokMasuk, PemakaianBarang
)
from .forms import (
    RegistrasiForm, RegistrasiDetailFormSet, PemasukanForm, PengeluaranForm, 
    ProgressTrackingForm, UserPasswordChangeForm, UserCreateForm, UserForm, 
    RoleForm, TemplatePesanForm, KategoriBarangForm, BarangInventoryForm, 
    StokMasukForm, PemakaianBarangForm
)
from django.db import transaction
from decimal import Decimal
from datetime import date
from core.services.notification_service import generate_all_notifications
from core.services.whatsapp_service import whatsapp_service
from .rbac import DEFAULT_ROLE_BLUEPRINTS, can_manage_roles, is_rbac_bootstrap_mode, seed_default_roles, sync_permission_catalog


def _expects_json_response(request):
    accept_header = request.headers.get('Accept', '')
    requested_with = request.headers.get('X-Requested-With')
    return (
        request.path.startswith('/api/')
        or request.path.startswith('/ajax/')
        or requested_with == 'XMLHttpRequest'
        or 'application/json' in accept_header
    )


def deny_permission_response(request, permission_code=None, message=None):
    message = message or 'Anda tidak punya akses untuk fitur ini.'
    if _expects_json_response(request):
        payload = {
            'success': False,
            'message': message,
        }
        if permission_code:
            payload['permission'] = permission_code
        return JsonResponse(payload, status=403)

    messages.error(request, message)
    return redirect('dashboard')


def require_permission(code, message=None):
    def decorator(view_func):
        @login_required
        @wraps(view_func)
        def _wrapped(request, *args, **kwargs):
            if getattr(request.user, 'has_permission', lambda c: False)(code):
                return view_func(request, *args, **kwargs)
            return deny_permission_response(request, code, message)

        return _wrapped

    return decorator


class PermissionRequiredViewMixin(LoginRequiredMixin):
    login_url = '/login/'
    permission_required = None
    permission_map = None
    permission_denied_message = 'Anda tidak punya akses untuk fitur ini.'

    def get_required_permission(self):
        if self.permission_map:
            return self.permission_map.get(self.request.method, self.permission_required)
        return self.permission_required

    def has_required_permission(self):
        permission_code = self.get_required_permission()
        if not permission_code:
            return True
        return getattr(self.request.user, 'has_permission', lambda c: False)(permission_code)

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()

        if not self.has_required_permission():
            return deny_permission_response(
                request,
                self.get_required_permission(),
                self.permission_denied_message,
            )

        return super().dispatch(request, *args, **kwargs)


def format_rupiah(value):
    """Format number as Indonesian Rupiah currency."""
    try:
        if isinstance(value, str):
            value = int(float(value))
        else:
            value = int(value)
        formatted = "{:,}".format(value).replace(",", ".")
        return f"Rp {formatted}"
    except (ValueError, TypeError):
        return "Rp 0"


class HealthCheckView(View):
    """Simple health check endpoint."""
    def get(self, request):
        return HttpResponse("OK")


@never_cache
def manifest_json(request):
    response = render(request, 'pwa/manifest.json', content_type='application/manifest+json')
    return response


@never_cache
def service_worker(request):
    response = render(request, 'pwa/sw.js', {
        'cache_version': 'babycare-v20260331-pwa2'
    }, content_type='application/javascript')
    response['Service-Worker-Allowed'] = '/'
    return response


@never_cache
def offline_view(request):
    return render(request, 'offline.html')


class DebugView(View):
    """Debug endpoint to check auth state."""
    def get(self, request):
        return HttpResponse(f"authenticated={request.user.is_authenticated}, user={request.user}")


class IndexView(View):
    """Redirect to login if not authenticated, dashboard if authenticated."""
    def get(self, request):
        if request.user.is_authenticated:
            return redirect('dashboard')
        return redirect('custom_login')


class SimpleLoginView(FormView):
    """Simple custom login view for debugging."""
    template_name = 'core/simple_login.html'
    form_class = AuthenticationForm
    success_url = reverse_lazy('dashboard')

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        user = form.get_user()
        login(self.request, user)
        return super().form_valid(form)
    
    def form_invalid(self, form):
        # Return form with errors
        return self.render_to_response(self.get_context_data(form=form))


class DashboardView(LoginRequiredMixin, View):
    """Simple dashboard view."""
    login_url = '/login/'
    
    def get(self, request):
        from django.shortcuts import render
        from django.utils import timezone
        from datetime import datetime, timedelta
        from django.db.models import Sum, Count, Q
        import json
        
        today = timezone.now().date()
        current_month_start = today.replace(day=1)
        
        # Calculate statistics
        total_pasien = Pasien.objects.count()
        registrasi_hari_ini = Registrasi.objects.filter(tanggal_kunjungan=today).count()
        pendapatan_bulan_ini = Pemasukan.objects.filter(
            tanggal__year=today.year,
            tanggal__month=today.month
        ).aggregate(total=Sum('jumlah'))['total'] or 0
        notifikasi_belum_dibaca = Notifikasi.objects.filter(sudah_dibaca=False).count()
        
        # Additional stats
        total_terapis_aktif = Terapis.objects.filter(is_active=True, is_deleted=False).count()
        registrasi_bulan_ini = Registrasi.objects.filter(
            tanggal_kunjungan__year=today.year,
            tanggal_kunjungan__month=today.month
        ).count()
        
        # Previous month comparison
        prev_month = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
        registrasi_bulan_lalu = Registrasi.objects.filter(
            tanggal_kunjungan__year=prev_month.year,
            tanggal_kunjungan__month=prev_month.month
        ).count()
        trend_registrasi = ((registrasi_bulan_ini - registrasi_bulan_lalu) / registrasi_bulan_lalu * 100) if registrasi_bulan_lalu > 0 else 0
        
        # Chart Data: Pendapatan Bulanan (6 bulan terakhir)
        pendapatan_per_bulan = []
        labels_pendapatan = []
        for i in range(5, -1, -1):
            month_date = today - timedelta(days=30*i)
            month_start = month_date.replace(day=1)
            if i > 0:
                next_month = (month_start + timedelta(days=32)).replace(day=1)
                month_end = next_month - timedelta(days=1)
            else:
                month_end = today
            
            revenue = Pemasukan.objects.filter(
                tanggal__gte=month_start,
                tanggal__lte=month_end
            ).aggregate(total=Sum('jumlah'))['total'] or 0
            pendapatan_per_bulan.append(float(revenue))
            labels_pendapatan.append(month_date.strftime('%b %Y'))
        
        # Chart Data 1: Registrasi per bulan (last 12 months)
        registrasi_per_bulan = []
        labels_bulan = []
        for i in range(11, -1, -1):
            month_date = today - timedelta(days=30*i)
            month_start = month_date.replace(day=1)
            if i > 0:
                month_end = (month_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            else:
                month_end = today
            
            count = Registrasi.objects.filter(
                tanggal_kunjungan__gte=month_start,
                tanggal_kunjungan__lte=month_end
            ).count()
            registrasi_per_bulan.append(count)
            labels_bulan.append(month_date.strftime('%b'))
        
        # Chart Data 2: Top Terapis (by number of sessions)
        top_terapis = Registrasi.objects.values('terapis__nama_terapis').annotate(
            count=Count('id')
        ).order_by('-count')[:8]
        terapis_names = [t['terapis__nama_terapis'] or 'Unknown' for t in top_terapis]
        terapis_counts = [t['count'] for t in top_terapis]
        
        # Chart Data 3: Top 5 Jenis Terapi (by frequency) - NEW
        top_jenis_terapi = RegistrasiDetail.objects.values('nama_terapi').annotate(
            count=Count('id'),
            revenue=Sum('harga_terapi')
        ).order_by('-count')[:5]
        top_terapi_names = [t['nama_terapi'] for t in top_jenis_terapi]
        top_terapi_counts = [t['count'] for t in top_jenis_terapi]
        
        # Chart Data 4: Jenis Terapi (by frequency) - Keep for backward compatibility
        jenis_terapi_data = Registrasi.objects.values('jenis_terapi__nama_terapi').annotate(
            count=Count('id')
        ).order_by('-count')[:8]
        terapi_names = [t['jenis_terapi__nama_terapi'] for t in jenis_terapi_data]
        terapi_counts = [t['count'] for t in jenis_terapi_data]
        
        # Chart Data 4: Pemasukan vs Pengeluaran (current month)
        pemasukan_bulan = Pemasukan.objects.filter(
            tanggal__year=today.year,
            tanggal__month=today.month
        ).aggregate(total=Sum('jumlah'))['total'] or 0
        
        pengeluaran_bulan = Pengeluaran.objects.filter(
            tanggal__year=today.year,
            tanggal__month=today.month
        ).aggregate(total=Sum('jumlah'))['total'] or 0
        
        # Chart Data 5: Pasien per cabang
        pasien_per_cabang = Pasien.objects.values('cabang__nama_cabang').annotate(
            count=Count('id')
        ).order_by('-count')
        cabang_names = [p['cabang__nama_cabang'] or 'Tanpa Cabang' for p in pasien_per_cabang]
        cabang_counts = [p['count'] for p in pasien_per_cabang]
        
        # Chart Data 6: Daily registrations trend (last 30 days)
        daily_registrasi = []
        labels_hari = []
        for i in range(29, -1, -1):
            day = today - timedelta(days=i)
            count = Registrasi.objects.filter(tanggal_kunjungan=day).count()
            daily_registrasi.append(count)
            labels_hari.append(day.strftime('%d'))
        
        context = {
            'username': request.user.username,
            'fullname': getattr(request.user, 'full_name', 'N/A'),
            'is_authenticated': request.user.is_authenticated,
            'total_pasien': total_pasien,
            'registrasi_hari_ini': registrasi_hari_ini,
            'pendapatan_bulan_ini': pendapatan_bulan_ini,
            'notifikasi_belum_dibaca': notifikasi_belum_dibaca,
            'total_terapis_aktif': total_terapis_aktif,
            'registrasi_bulan_ini': registrasi_bulan_ini,
            'trend_registrasi': round(trend_registrasi, 1),
            # Chart data for revenue (6 months)
            'pendapatan_per_bulan': json.dumps(pendapatan_per_bulan),
            'labels_pendapatan': json.dumps(labels_pendapatan),
            # Chart data for top 5 therapy types
            'top_terapi_names': json.dumps(top_terapi_names),
            'top_terapi_counts': json.dumps(top_terapi_counts),
            # Original chart data (keep for compatibility)
            'registrasi_per_bulan': json.dumps(registrasi_per_bulan),
            'labels_bulan': json.dumps(labels_bulan),
            'terapis_names': json.dumps(terapis_names),
            'terapis_counts': json.dumps(terapis_counts),
            'terapi_names': json.dumps(terapi_names),
            'terapi_counts': json.dumps(terapi_counts),
            'pemasukan_bulan': float(pemasukan_bulan),
            'pengeluaran_bulan': float(pengeluaran_bulan),
            'cabang_names': json.dumps(cabang_names),
            'cabang_counts': json.dumps(cabang_counts),
            'daily_registrasi': json.dumps(daily_registrasi),
            'labels_hari': json.dumps(labels_hari),
        }
        return render(request, 'core/dashboard.html', context)

class RegistrasiListView(PermissionRequiredViewMixin, ListView):
    model = Registrasi
    template_name = 'core/registrasi_list.html'
    context_object_name = 'registrasis'
    paginate_by = 50
    permission_required = 'registrasi_view'

    def get_queryset(self):
        qs = super().get_queryset().select_related(
            'pasien', 'jenis_terapi', 'terapis', 'cabang'
        ).order_by('-tanggal_kunjungan')
        
        # Add annotation for outstanding balance
        from django.db.models import Sum, F, DecimalField, Case, When
        from decimal import Decimal
        qs = qs.annotate(
            total_paid=Sum('pemasukan__jumlah', output_field=DecimalField())
        )
        # Calculate remaining balance as total_bayar - total_paid (default 0 if null)
        qs = qs.annotate(
            sisa_tagihan=Case(
                When(total_paid__isnull=True, then=F('total_bayar')),
                default=F('total_bayar') - F('total_paid'),
                output_field=DecimalField()
            )
        )
        
        user_roles = getattr(self.request, 'user_roles', set())
        if 'terapis' in user_roles:
            # Terapis model in the active schema has no direct FK to User.
            # Fall back to cabang scoping to avoid invalid joins.
            if getattr(self.request, 'cabang_id', None) is not None:
                qs = qs.filter(cabang_id=self.request.cabang_id)
        else:
            if getattr(self.request, 'cabang_id', None) is not None:
                qs = qs.filter(cabang_id=self.request.cabang_id)
        
        # Apply filters from GET parameters
        # Filter by date range
        tanggal_dari = self.request.GET.get('tanggal_dari')
        tanggal_sampai = self.request.GET.get('tanggal_sampai')
        
        if tanggal_dari:
            try:
                from datetime import datetime
                tanggal_dari_date = datetime.strptime(tanggal_dari, '%Y-%m-%d').date()
                qs = qs.filter(tanggal_kunjungan__gte=tanggal_dari_date)
            except:
                pass
        
        if tanggal_sampai:
            try:
                from datetime import datetime
                tanggal_sampai_date = datetime.strptime(tanggal_sampai, '%Y-%m-%d').date()
                qs = qs.filter(tanggal_kunjungan__lte=tanggal_sampai_date)
            except:
                pass
        
        # Filter by pasien (nama_anak)
        pasien_query = self.request.GET.get('pasien_query')
        if pasien_query:
            qs = qs.filter(pasien__nama_anak__icontains=pasien_query)
        
        # Filter by jenis_terapi
        jenis_terapi_id = self.request.GET.get('jenis_terapi_id')
        if jenis_terapi_id:
            qs = qs.filter(jenis_terapi_id=jenis_terapi_id)
        
        # Filter by terapis
        terapis_id = self.request.GET.get('terapis_id')
        if terapis_id:
            qs = qs.filter(terapis_id=terapis_id)
        
        # Filter by kode_registrasi
        kode_registrasi = self.request.GET.get('kode_registrasi')
        if kode_registrasi:
            qs = qs.filter(kode_registrasi__icontains=kode_registrasi)
        
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filter values
        context['tanggal_dari'] = self.request.GET.get('tanggal_dari', '')
        context['tanggal_sampai'] = self.request.GET.get('tanggal_sampai', '')
        context['pasien_query'] = self.request.GET.get('pasien_query', '')
        context['jenis_terapi_id'] = self.request.GET.get('jenis_terapi_id', '')
        context['terapis_id'] = self.request.GET.get('terapis_id', '')
        context['kode_registrasi'] = self.request.GET.get('kode_registrasi', '')
        
        # Get filter options
        from core.models import JenisTerapi, Terapis
        context['jenis_terapi_list'] = JenisTerapi.objects.all()
        context['terapis_list'] = Terapis.objects.all()
        context['whatsapp_configured'] = whatsapp_service.is_configured()
        
        return context

class RegistrasiCreateView(PermissionRequiredViewMixin, CreateView):
    model = Registrasi
    form_class = RegistrasiForm
    template_name = 'core/registrasi_form.html'
    success_url = reverse_lazy('registrasi_list')
    permission_required = 'registrasi_create'

    def get_initial(self):
        """Set default values for form fields."""
        initial = super().get_initial()
        initial['tanggal_kunjungan'] = date.today()
        return initial

    def get_context_data(self, **kwargs):
        """Add formset to context"""
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = RegistrasiDetailFormSet(self.request.POST)
        else:
            context['formset'] = RegistrasiDetailFormSet()
        context['is_edit'] = False
        return context

    @transaction.atomic
    def form_valid(self, form):
        """Handle form and formset validation & saving"""
        context = self.get_context_data()
        formset = context['formset']
        
        # Validate formset
        if not formset.is_valid():
            print("Formset errors:", formset.errors)
            messages.error(self.request, 'Terjadi kesalahan pada data terapi. Periksa kembali input Anda.')
            return self.form_invalid(form)
        
        # Generate kode_registrasi
        if not form.instance.kode_registrasi:
            today = date.today()
            cabang_id = form.instance.cabang_id
            cabang_code = f'{cabang_id:02d}' if cabang_id else '00'
            month_year = today.strftime('%m%y')
            prefix = f'P{cabang_code}{month_year}'
            
            last_registrasi = Registrasi.objects.filter(
                kode_registrasi__startswith=prefix,
                cabang_id=cabang_id
            ).order_by('-kode_registrasi', '-id').first()
            
            if last_registrasi and last_registrasi.kode_registrasi:
                try:
                    last_seq = int(last_registrasi.kode_registrasi[-3:])
                    next_seq = last_seq + 1
                except (ValueError, IndexError):
                    next_seq = 1
            else:
                next_seq = 1
            
            form.instance.kode_registrasi = f'{prefix}{next_seq:03d}'
        
        # Calculate total harga from all terapi in formset
        total_harga = Decimal('0.00')
        first_terapi = None
        for detail_form in formset:
            if detail_form.cleaned_data and not detail_form.cleaned_data.get('DELETE'):
                jenis_terapi = detail_form.cleaned_data.get('jenis_terapi')
                if jenis_terapi:
                    if first_terapi is None:
                        first_terapi = jenis_terapi
                    total_harga += jenis_terapi.harga
        
        # Set jenis_terapi (use first therapy) and harga
        if first_terapi:
            form.instance.jenis_terapi = first_terapi
        form.instance.harga = total_harga
        biaya_transport = form.instance.biaya_transport or Decimal('0.00')
        form.instance.total_bayar = total_harga + biaya_transport
        
        # Save registrasi (header)
        self.object = form.save()
        
        # Save detail terapi
        formset.instance = self.object
        details = formset.save(commit=False)
        
        for detail in details:
            detail.registrasi = self.object
            detail.kode_registrasi = self.object.kode_registrasi
            detail.nama_terapi = detail.jenis_terapi.nama_terapi
            detail.harga_terapi = detail.jenis_terapi.harga
            detail.save()
        
        # Delete any marked for deletion
        for obj in formset.deleted_objects:
            obj.delete()
        
        messages.success(self.request, f'Registrasi berhasil disimpan! Kode: {self.object.kode_registrasi}')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        print("Form errors:", form.errors)
        error_messages = []
        for field, errors in form.errors.items():
            if field == '__all__':
                error_messages.extend(errors)
            else:
                field_label = form.fields.get(field, type('obj', (), {'label': field})).label or field
                for error in errors:
                    error_messages.append(f"{field_label}: {error}")
        
        if error_messages:
            messages.error(self.request, 'Terjadi kesalahan:<br>' + '<br>'.join(error_messages))
        else:
            messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        
        return super().form_invalid(form)

class RegistrasiEditView(PermissionRequiredViewMixin, UpdateView):
    model = Registrasi
    form_class = RegistrasiForm
    template_name = 'core/registrasi_form.html'
    success_url = reverse_lazy('registrasi_list')
    permission_required = 'registrasi_edit'

    def get_context_data(self, **kwargs):
        """Add formset to context"""
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = RegistrasiDetailFormSet(self.request.POST, instance=self.object)
        else:
            context['formset'] = RegistrasiDetailFormSet(instance=self.object)
        context['is_edit'] = True
        context['progress_form'] = ProgressTrackingForm()
        context['progress_entries'] = self.object.progress_entries.select_related('created_by').all()
        return context

    @transaction.atomic
    def form_valid(self, form):
        """Handle form and formset validation & saving"""
        context = self.get_context_data()
        formset = context['formset']
        
        # Validate formset
        if not formset.is_valid():
            print("Formset errors:", formset.errors)
            messages.error(self.request, 'Terjadi kesalahan pada data terapi. Periksa kembali input Anda.')
            return self.form_invalid(form)
        
        # Calculate total harga from all terapi in formset
        total_harga = Decimal('0.00')
        first_terapi = None
        for detail_form in formset:
            if detail_form.cleaned_data and not detail_form.cleaned_data.get('DELETE'):
                jenis_terapi = detail_form.cleaned_data.get('jenis_terapi')
                if jenis_terapi:
                    if first_terapi is None:
                        first_terapi = jenis_terapi
                    total_harga += jenis_terapi.harga
        
        # Update jenis_terapi (use first therapy) and harga
        if first_terapi:
            form.instance.jenis_terapi = first_terapi
        form.instance.harga = total_harga
        biaya_transport = form.instance.biaya_transport or Decimal('0.00')
        form.instance.total_bayar = total_harga + biaya_transport
        
        # Save registrasi (header)
        self.object = form.save()
        
        # Save detail terapi
        formset.instance = self.object
        details = formset.save(commit=False)
        
        for detail in details:
            detail.registrasi = self.object
            detail.kode_registrasi = self.object.kode_registrasi
            detail.nama_terapi = detail.jenis_terapi.nama_terapi
            detail.harga_terapi = detail.jenis_terapi.harga
            detail.save()
        
        # Delete any marked for deletion
        for obj in formset.deleted_objects:
            obj.delete()
        
        messages.success(self.request, f'Registrasi berhasil diupdate! Kode: {self.object.kode_registrasi}')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        error_messages = []
        for field, errors in form.errors.items():
            if field == '__all__':
                error_messages.extend(errors)
            else:
                field_label = form.fields.get(field, type('obj', (), {'label': field})).label or field
                for error in errors:
                    error_messages.append(f"{field_label}: {error}")
        
        if error_messages:
            messages.error(self.request, 'Terjadi kesalahan:<br>' + '<br>'.join(error_messages))
        else:
            messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        
        return super().form_invalid(form)


@require_permission('registrasi_send_whatsapp')
@require_http_methods(["POST"])
def send_whatsapp_reminder(request, registrasi_id):
    registrasi = get_object_or_404(
        Registrasi.objects.select_related('pasien', 'terapis', 'cabang').prefetch_related('details__jenis_terapi'),
        pk=registrasi_id,
        is_deleted=False,
    )

    result = whatsapp_service.send_appointment_reminder(registrasi)
    status_code = 200 if result['success'] else 400
    return JsonResponse(result, status=status_code)


@require_permission('registrasi_add_progress')
@require_http_methods(["POST"])
def add_progress_tracking(request, registrasi_id):
    registrasi = get_object_or_404(Registrasi, pk=registrasi_id, is_deleted=False)
    form = ProgressTrackingForm(request.POST, request.FILES)

    if form.is_valid():
        progress = form.save(commit=False)
        progress.registrasi = registrasi
        progress.created_by = request.user
        progress.save()
        messages.success(request, 'Progress tracking berhasil ditambahkan.')
    else:
        error_text = ' '.join([' '.join(errors) for errors in form.errors.values()])
        messages.error(request, error_text or 'Gagal menambahkan progress tracking.')

    return redirect('registrasi_edit', pk=registrasi_id)

class PemasukanCreateView(PermissionRequiredViewMixin, CreateView):
    model = Pemasukan
    form_class = PemasukanForm
    template_name = 'core/pemasukan_form.html'
    success_url = reverse_lazy('pemasukan_list')
    permission_required = 'pemasukan_create'

    def get_success_url(self):
        """Check if there's a 'next' parameter and use it, otherwise use default."""
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return self.success_url

    def get_initial(self):
        """Set default values for form fields."""
        initial = super().get_initial()
        from datetime import date
        initial['tanggal'] = date.today()
        return initial

    def get_form(self, form_class=None):
        """Override to filter registrasi queryset - only show unpaid/partially paid registrations."""
        form = super().get_form(form_class)
        
        # Get all registrasi IDs with their total payments
        from django.db.models import Sum, Q
        paid_registrasi = Pemasukan.objects.values('registrasi_id').annotate(
            total_paid=Sum('jumlah')
        )
        
        # Build dict of registrasi_id -> total_paid
        paid_dict = {item['registrasi_id']: item['total_paid'] for item in paid_registrasi if item['registrasi_id']}
        
        # Filter registrasi: exclude fully paid ones
        unpaid_registrasi = []
        for reg in Registrasi.objects.filter(is_deleted=False).select_related('pasien', 'jenis_terapi', 'terapis'):
            total_paid = paid_dict.get(reg.id, 0) or 0
            total_bayar = reg.total_bayar or 0
            
            # Include if not fully paid
            if total_paid < total_bayar:
                unpaid_registrasi.append(reg.id)
        
        # Update form queryset
        form.fields['registrasi'].queryset = Registrasi.objects.filter(
            id__in=unpaid_registrasi
        ).select_related('pasien', 'jenis_terapi', 'terapis').order_by('-tanggal_kunjungan')
        
        return form

    def form_valid(self, form):
        try:
            obj = form.save(commit=False)
            if obj is None:
                print("ERROR: form.save(commit=False) returned None!")
                messages.error(self.request, 'Terjadi kesalahan saat menyimpan data.')
                return self.form_invalid(form)
            obj.created_by = self.request.user
            obj.save()
            self.object = obj  # IMPORTANT: Set self.object for get_success_url()
            
            # Create success message with kembalian info if applicable
            msg = '✅ Pembayaran berhasil disimpan!'
            if obj.jumlah_bayar and obj.jumlah and obj.jumlah_bayar > obj.jumlah:
                kembalian = obj.jumlah_bayar - obj.jumlah
                msg += f' Kembalian: Rp {kembalian:,.0f}'.replace(',', '.')
            
            messages.success(self.request, msg)
            return redirect(self.get_success_url())
        except AttributeError as e:
            print(f"AttributeError in form_valid: {e}")
            print(f"form.save() returned: {form.save(commit=False) if hasattr(form, 'save') else 'No save method'}")
            messages.error(self.request, f'Terjadi kesalahan teknis: {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            print(f"Exception in form_valid: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            messages.error(self.request, f'Terjadi kesalahan: {str(e)}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        # Debug: print form errors to console
        print("Form errors:", form.errors)
        print("Form cleaned_data:", getattr(form, 'cleaned_data', 'No cleaned data'))
        
        # Create detailed error message
        error_messages = []
        for field, errors in form.errors.items():
            if field == '__all__':
                error_messages.extend(errors)
            else:
                field_label = form.fields[field].label or field
                for error in errors:
                    error_messages.append(f"{field_label}: {error}")
        
        if error_messages:
            messages.error(self.request, 'Terjadi kesalahan:<br>' + '<br>'.join(error_messages))
        else:
            messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        
        return super().form_invalid(form)


class PemasukanListView(PermissionRequiredViewMixin, ListView):
    model = Pemasukan
    template_name = 'core/pemasukan_list.html'
    context_object_name = 'pemasukans'
    paginate_by = 25
    permission_required = 'pemasukan_view'

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format == 'excel':
            return self.export_to_excel(request)
        if export_format == 'pdf':
            return self.export_to_pdf(request)
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        from datetime import datetime

        qs = super().get_queryset().select_related('registrasi__pasien', 'registrasi__jenis_terapi', 'cabang').order_by('-tanggal', '-created_at')
        if getattr(self.request, 'cabang_id', None) is not None:
            qs = qs.filter(cabang_id=self.request.cabang_id)

        tanggal_dari = self.request.GET.get('tanggal_dari')
        if tanggal_dari:
            try:
                qs = qs.filter(tanggal__gte=datetime.strptime(tanggal_dari, '%Y-%m-%d').date())
            except ValueError:
                pass

        tanggal_sampai = self.request.GET.get('tanggal_sampai')
        if tanggal_sampai:
            try:
                qs = qs.filter(tanggal__lte=datetime.strptime(tanggal_sampai, '%Y-%m-%d').date())
            except ValueError:
                pass

        cabang_id = self.request.GET.get('cabang_id')
        if cabang_id and getattr(self.request, 'cabang_id', None) is None:
            qs = qs.filter(cabang_id=cabang_id)

        metode_pembayaran = self.request.GET.get('metode_pembayaran')
        if metode_pembayaran:
            qs = qs.filter(metode_pembayaran=metode_pembayaran)

        keyword = (self.request.GET.get('keyword') or '').strip()
        if keyword:
            qs = qs.filter(
                Q(keterangan__icontains=keyword)
                | Q(registrasi__pasien__nama_anak__icontains=keyword)
                | Q(registrasi__kode_registrasi__icontains=keyword)
            )
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.db.models import Sum
        # Calculate total from all records (not just current page)
        qs = self.get_queryset()
        total = qs.aggregate(total=Sum('jumlah'))['total'] or 0
        context['total_pemasukan'] = total
        context['tanggal_dari'] = self.request.GET.get('tanggal_dari', '')
        context['tanggal_sampai'] = self.request.GET.get('tanggal_sampai', '')
        context['cabang_id'] = self.request.GET.get('cabang_id', '')
        context['metode_pembayaran'] = self.request.GET.get('metode_pembayaran', '')
        context['keyword'] = self.request.GET.get('keyword', '')
        context['cabang_list'] = Cabang.objects.all().order_by('nama_cabang')
        context['metode_choices'] = (
            ('TUNAI', 'Tunai'),
            ('TRANSFER', 'Transfer'),
            ('QRIS', 'QRIS'),
            ('DEBIT', 'Debit'),
            ('KREDIT', 'Kredit'),
        )
        return context

    def export_to_excel(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        queryset = self.get_queryset()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Pemasukan'

        sheet.column_dimensions['A'].width = 14
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 28
        sheet.column_dimensions['D'].width = 28
        sheet.column_dimensions['E'].width = 18
        sheet.column_dimensions['F'].width = 16
        sheet.column_dimensions['G'].width = 22
        sheet.column_dimensions['H'].width = 36

        header_fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )

        sheet.merge_cells('A1:H1')
        sheet['A1'] = 'LAPORAN PEMASUKAN'
        sheet['A1'].font = Font(size=14, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')

        sheet.merge_cells('A2:H2')
        sheet['A2'] = f"Diekspor: {timezone.localtime().strftime('%d %B %Y %H:%M')}"
        sheet['A2'].alignment = Alignment(horizontal='center')

        headers = ['Tanggal', 'Kode', 'Pasien', 'Cabang', 'Jumlah', 'Metode', 'Kasir', 'Keterangan']
        for index, label in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=index, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        for row_index, item in enumerate(queryset, start=5):
            pasien = item.registrasi.pasien.nama_anak if item.registrasi and item.registrasi.pasien else '-'
            kode = item.registrasi.kode_registrasi if item.registrasi else '-'
            kasir = item.created_by.full_name if item.created_by and item.created_by.full_name else (item.created_by.username if item.created_by else '-')
            row_values = [
                item.tanggal.strftime('%d/%m/%Y') if item.tanggal else '-',
                kode,
                pasien,
                item.cabang.nama_cabang if item.cabang else '-',
                float(item.jumlah or 0),
                item.metode_pembayaran or '-',
                kasir,
                item.keterangan or '-',
            ]
            for column_index, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_index, column=column_index, value=value)
                cell.border = border
                if column_index == 5:
                    cell.number_format = '#,##0'

        total_row = queryset.count() + 5
        sheet.cell(row=total_row, column=1, value='Total').font = Font(bold=True)
        sheet.cell(row=total_row, column=5, value=float(queryset.aggregate(total=Sum('jumlah'))['total'] or 0)).font = Font(bold=True)
        sheet.cell(row=total_row, column=5).number_format = '#,##0'

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="pemasukan_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        workbook.save(response)
        return response

    def export_to_pdf(self, request):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        queryset = self.get_queryset()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="pemasukan_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

        document = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
        styles = getSampleStyleSheet()
        elements = [
            Paragraph('Laporan Pemasukan', styles['Title']),
            Paragraph(f"Diekspor: {timezone.localtime().strftime('%d %B %Y %H:%M')}", styles['Normal']),
            Spacer(1, 10),
        ]

        data = [[
            'Tanggal', 'Kode', 'Pasien', 'Cabang', 'Jumlah', 'Metode', 'Kasir', 'Keterangan'
        ]]
        for item in queryset:
            pasien = item.registrasi.pasien.nama_anak if item.registrasi and item.registrasi.pasien else '-'
            kode = item.registrasi.kode_registrasi if item.registrasi else '-'
            kasir = item.created_by.full_name if item.created_by and item.created_by.full_name else (item.created_by.username if item.created_by else '-')
            data.append([
                item.tanggal.strftime('%d/%m/%Y') if item.tanggal else '-',
                kode,
                pasien,
                item.cabang.nama_cabang if item.cabang else '-',
                format_rupiah(item.jumlah or 0),
                item.metode_pembayaran or '-',
                kasir,
                (item.keterangan or '-')[:70],
            ])

        data.append(['', '', '', 'Total', format_rupiah(queryset.aggregate(total=Sum('jumlah'))['total'] or 0), '', '', ''])

        table = Table(data, repeatRows=1, colWidths=[22 * mm, 26 * mm, 42 * mm, 32 * mm, 26 * mm, 24 * mm, 30 * mm, 58 * mm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#198754')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ECFDF3')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)
        document.build(elements)
        return response


class PemasukanEditView(PermissionRequiredViewMixin, UpdateView):
    model = Pemasukan
    form_class = PemasukanForm
    template_name = 'core/pemasukan_form.html'
    success_url = reverse_lazy('pemasukan_list')
    permission_required = 'pemasukan_edit'

    def get_form(self, form_class=None):
        """Override to filter registrasi queryset - only show unpaid/partially paid registrations."""
        form = super().get_form(form_class)
        
        # Get all registrasi IDs with their total payments (excluding current pemasukan being edited)
        from django.db.models import Sum, Q
        paid_registrasi = Pemasukan.objects.exclude(pk=self.object.pk).values('registrasi_id').annotate(
            total_paid=Sum('jumlah')
        )
        
        # Build dict of registrasi_id -> total_paid
        paid_dict = {item['registrasi_id']: item['total_paid'] for item in paid_registrasi if item['registrasi_id']}
        
        # Filter registrasi: exclude fully paid ones
        unpaid_registrasi = []
        for reg in Registrasi.objects.filter(is_deleted=False).select_related('pasien', 'jenis_terapi', 'terapis'):
            total_paid = paid_dict.get(reg.id, 0) or 0
            total_bayar = reg.total_bayar or 0
            
            # Include if not fully paid OR if it's the current pemasukan's registrasi
            if total_paid < total_bayar or reg.id == (self.object.registrasi_id if self.object.registrasi else None):
                unpaid_registrasi.append(reg.id)
        
        # Update form queryset
        form.fields['registrasi'].queryset = Registrasi.objects.filter(
            id__in=unpaid_registrasi
        ).select_related('pasien', 'jenis_terapi', 'terapis').order_by('-tanggal_kunjungan')
        
        return form

    def form_valid(self, form):
        obj = form.save()
        
        # Create success message with kembalian info if applicable
        msg = '✅ Data pembayaran berhasil diupdate!'
        if obj.jumlah_bayar and obj.jumlah and obj.jumlah_bayar > obj.jumlah:
            kembalian = obj.jumlah_bayar - obj.jumlah
            msg += f' Kembalian: Rp {kembalian:,.0f}'.replace(',', '.')
        
        messages.success(self.request, msg)
        return super().form_valid(form)
    
    def form_invalid(self, form):
        error_messages = []
        for field, errors in form.errors.items():
            if field == '__all__':
                error_messages.extend(errors)
            else:
                field_label = form.fields[field].label or field
                for error in errors:
                    error_messages.append(f"{field_label}: {error}")
        
        if error_messages:
            messages.error(self.request, 'Terjadi kesalahan:<br>' + '<br>'.join(error_messages))
        else:
            messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        
        return super().form_invalid(form)


class PengeluaranListView(PermissionRequiredViewMixin, ListView):
    model = Pengeluaran
    template_name = 'core/pengeluaran_list.html'
    context_object_name = 'pengeluarans'
    paginate_by = 25
    permission_required = 'pengeluaran_view'

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format == 'excel':
            return self.export_to_excel(request)
        if export_format == 'pdf':
            return self.export_to_pdf(request)
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        from datetime import datetime

        qs = super().get_queryset().order_by('-tanggal', '-created_at')
        if getattr(self.request, 'cabang_id', None) is not None:
            qs = qs.filter(cabang_id=self.request.cabang_id)

        tanggal_dari = self.request.GET.get('tanggal_dari')
        if tanggal_dari:
            try:
                qs = qs.filter(tanggal__gte=datetime.strptime(tanggal_dari, '%Y-%m-%d').date())
            except ValueError:
                pass

        tanggal_sampai = self.request.GET.get('tanggal_sampai')
        if tanggal_sampai:
            try:
                qs = qs.filter(tanggal__lte=datetime.strptime(tanggal_sampai, '%Y-%m-%d').date())
            except ValueError:
                pass

        cabang_id = self.request.GET.get('cabang_id')
        if cabang_id and getattr(self.request, 'cabang_id', None) is None:
            qs = qs.filter(cabang_id=cabang_id)

        kategori = (self.request.GET.get('kategori') or '').strip()
        if kategori:
            qs = qs.filter(kategori=kategori)

        keyword = (self.request.GET.get('keyword') or '').strip()
        if keyword:
            qs = qs.filter(Q(keterangan__icontains=keyword) | Q(kategori__icontains=keyword))
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.db.models import Sum
        # Calculate total from all records (not just current page)
        qs = self.get_queryset()
        total = qs.aggregate(total=Sum('jumlah'))['total'] or 0
        context['total_pengeluaran'] = total
        context['tanggal_dari'] = self.request.GET.get('tanggal_dari', '')
        context['tanggal_sampai'] = self.request.GET.get('tanggal_sampai', '')
        context['cabang_id'] = self.request.GET.get('cabang_id', '')
        context['kategori'] = self.request.GET.get('kategori', '')
        context['keyword'] = self.request.GET.get('keyword', '')
        context['cabang_list'] = Cabang.objects.all().order_by('nama_cabang')
        context['kategori_choices'] = Pengeluaran.objects.exclude(kategori__isnull=True).exclude(kategori__exact='').values_list('kategori', flat=True).distinct().order_by('kategori')
        return context

    def export_to_excel(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        queryset = self.get_queryset().select_related('cabang', 'created_by')
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Pengeluaran'

        sheet.column_dimensions['A'].width = 14
        sheet.column_dimensions['B'].width = 24
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 18
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 40

        header_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )

        sheet.merge_cells('A1:F1')
        sheet['A1'] = 'LAPORAN PENGELUARAN'
        sheet['A1'].font = Font(size=14, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')

        sheet.merge_cells('A2:F2')
        sheet['A2'] = f"Diekspor: {timezone.localtime().strftime('%d %B %Y %H:%M')}"
        sheet['A2'].alignment = Alignment(horizontal='center')

        headers = ['Tanggal', 'Kategori', 'Cabang', 'Jumlah', 'Dibuat Oleh', 'Keterangan']
        for index, label in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=index, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        for row_index, item in enumerate(queryset, start=5):
            kasir = item.created_by.full_name if item.created_by and item.created_by.full_name else (item.created_by.username if item.created_by else '-')
            row_values = [
                item.tanggal.strftime('%d/%m/%Y') if item.tanggal else '-',
                item.kategori or '-',
                item.cabang.nama_cabang if item.cabang else '-',
                float(item.jumlah or 0),
                kasir,
                item.keterangan or '-',
            ]
            for column_index, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_index, column=column_index, value=value)
                cell.border = border
                if column_index == 4:
                    cell.number_format = '#,##0'

        total_row = queryset.count() + 5
        sheet.cell(row=total_row, column=1, value='Total').font = Font(bold=True)
        sheet.cell(row=total_row, column=4, value=float(queryset.aggregate(total=Sum('jumlah'))['total'] or 0)).font = Font(bold=True)
        sheet.cell(row=total_row, column=4).number_format = '#,##0'

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="pengeluaran_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        workbook.save(response)
        return response

    def export_to_pdf(self, request):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        queryset = self.get_queryset().select_related('cabang', 'created_by')
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="pengeluaran_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

        document = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
        styles = getSampleStyleSheet()
        elements = [
            Paragraph('Laporan Pengeluaran', styles['Title']),
            Paragraph(f"Diekspor: {timezone.localtime().strftime('%d %B %Y %H:%M')}", styles['Normal']),
            Spacer(1, 10),
        ]

        data = [['Tanggal', 'Kategori', 'Cabang', 'Jumlah', 'Dibuat Oleh', 'Keterangan']]
        for item in queryset:
            kasir = item.created_by.full_name if item.created_by and item.created_by.full_name else (item.created_by.username if item.created_by else '-')
            data.append([
                item.tanggal.strftime('%d/%m/%Y') if item.tanggal else '-',
                item.kategori or '-',
                item.cabang.nama_cabang if item.cabang else '-',
                format_rupiah(item.jumlah or 0),
                kasir,
                (item.keterangan or '-')[:90],
            ])

        data.append(['', '', 'Total', format_rupiah(queryset.aggregate(total=Sum('jumlah'))['total'] or 0), '', ''])

        table = Table(data, repeatRows=1, colWidths=[26 * mm, 38 * mm, 34 * mm, 28 * mm, 38 * mm, 84 * mm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC3545')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF2F2')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)
        document.build(elements)
        return response


class PengeluaranCreateView(PermissionRequiredViewMixin, CreateView):
    model = Pengeluaran
    form_class = PengeluaranForm
    template_name = 'core/pengeluaran_form.html'
    success_url = reverse_lazy('pengeluaran_list')
    permission_required = 'pengeluaran_create'

    def get_initial(self):
        """Set default values for form fields."""
        initial = super().get_initial()
        from datetime import date
        initial['tanggal'] = date.today()
        return initial

    def form_valid(self, form):
        try:
            obj = form.save(commit=False)
            if obj is None:
                print("ERROR: form.save(commit=False) returned None!")
                messages.error(self.request, 'Terjadi kesalahan saat menyimpan data.')
                return self.form_invalid(form)
            obj.created_by = self.request.user
            obj.save()
            self.object = obj
            
            # Jika pembelian barang inventory, create StokMasuk otomatis
            if obj.barang and obj.jumlah_barang and obj.harga_satuan_beli:
                from .models import StokMasuk
                stok_masuk = StokMasuk.objects.create(
                    barang=obj.barang,
                    tanggal_masuk=obj.tanggal,
                    jumlah=obj.jumlah_barang,
                    harga_beli_satuan=obj.harga_satuan_beli,
                    supplier=obj.supplier or '-',
                    no_faktur=obj.no_faktur or '-',
                    cabang=obj.cabang,
                    catatan=f'Auto dari Pengeluaran ID: {obj.id}',
                    created_by=self.request.user
                )
                messages.success(self.request, f'Data pengeluaran berhasil disimpan! Stok {obj.barang.nama_barang} bertambah {obj.jumlah_barang} {obj.barang.satuan}.')
            else:
                messages.success(self.request, 'Data pengeluaran berhasil disimpan!')
            
            return redirect(self.get_success_url())
        except Exception as e:
            print(f"Exception in form_valid: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            messages.error(self.request, f'Terjadi kesalahan: {str(e)}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        # Debug: print form errors to console
        print("Form errors:", form.errors)
        print("Form cleaned_data:", getattr(form, 'cleaned_data', 'No cleaned data'))
        
        # Create detailed error message
        error_messages = []
        for field, errors in form.errors.items():
            if field == '__all__':
                error_messages.extend(errors)
            else:
                field_label = form.fields[field].label or field
                for error in errors:
                    error_messages.append(f"{field_label}: {error}")
        
        if error_messages:
            messages.error(self.request, 'Terjadi kesalahan:<br>' + '<br>'.join(error_messages))
        else:
            messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        
        return super().form_invalid(form)


class PengeluaranEditView(PermissionRequiredViewMixin, UpdateView):
    model = Pengeluaran
    form_class = PengeluaranForm
    template_name = 'core/pengeluaran_form.html'
    success_url = reverse_lazy('pengeluaran_list')
    permission_required = 'pengeluaran_edit'

    def form_valid(self, form):
        obj = form.save()
        
        # Jika pembelian barang inventory, create StokMasuk otomatis
        if obj.barang and obj.jumlah_barang and obj.harga_satuan_beli:
            from .models import StokMasuk
            # Cek apakah sudah ada stok masuk dari pengeluaran ini
            existing = StokMasuk.objects.filter(
                catatan__contains=f'Auto dari Pengeluaran ID: {obj.id}'
            ).first()
            
            if existing:
                # Update existing
                existing.barang = obj.barang
                existing.tanggal_masuk = obj.tanggal
                existing.jumlah = obj.jumlah_barang
                existing.harga_beli_satuan = obj.harga_satuan_beli
                existing.supplier = obj.supplier or '-'
                existing.no_faktur = obj.no_faktur or '-'
                existing.cabang = obj.cabang
                existing.save()
                messages.success(self.request, f'✅ Data pengeluaran berhasil diupdate! Stok {obj.barang.nama_barang} diupdate.')
            else:
                # Create new
                StokMasuk.objects.create(
                    barang=obj.barang,
                    tanggal_masuk=obj.tanggal,
                    jumlah=obj.jumlah_barang,
                    harga_beli_satuan=obj.harga_satuan_beli,
                    supplier=obj.supplier or '-',
                    no_faktur=obj.no_faktur or '-',
                    cabang=obj.cabang,
                    catatan=f'Auto dari Pengeluaran ID: {obj.id}',
                    created_by=self.request.user
                )
                messages.success(self.request, f'✅ Data pengeluaran berhasil diupdate! Stok {obj.barang.nama_barang} bertambah {obj.jumlah_barang} {obj.barang.satuan}.')
        else:
            messages.success(self.request, '✅ Data pengeluaran berhasil diupdate!')
        
        return super().form_valid(form)
    
    def form_invalid(self, form):
        error_messages = []
        for field, errors in form.errors.items():
            if field == '__all__':
                error_messages.extend(errors)
            else:
                field_label = form.fields[field].label or field
                for error in errors:
                    error_messages.append(f"{field_label}: {error}")
        
        if error_messages:
            messages.error(self.request, 'Terjadi kesalahan:<br>' + '<br>'.join(error_messages))
        else:
            messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        
        return super().form_invalid(form)


# Master Data Views
from .models import Pasien, JenisTerapi, Terapis

class PasienListView(PermissionRequiredViewMixin, ListView):
    """List all pasien."""
    model = Pasien
    template_name = 'core/pasien_list.html'
    context_object_name = 'pasiens'
    paginate_by = 25
    permission_required = 'pasien_view'
    
    def get_queryset(self):
        qs = super().get_queryset().select_related('cabang').order_by('-id')
        if getattr(self.request, 'cabang_id', None) is not None:
            qs = qs.filter(cabang_id=self.request.cabang_id)
        
        # Search by nama_anak or nama_orang_tua or no_wa
        search_query = self.request.GET.get('search')
        if search_query:
            from django.db.models import Q
            qs = qs.filter(
                Q(nama_anak__icontains=search_query) |
                Q(nama_orang_tua__icontains=search_query) |
                Q(no_wa__icontains=search_query) |
                Q(kode_pasien__icontains=search_query)
            )
        
        # Filter by cabang
        cabang_id = self.request.GET.get('cabang_id')
        if cabang_id:
            qs = qs.filter(cabang_id=cabang_id)
        
        # Filter by WhatsApp status
        has_wa = self.request.GET.get('has_whatsapp')
        if has_wa == 'true':
            qs = qs.filter(has_whatsapp=True)
        elif has_wa == 'false':
            qs = qs.filter(has_whatsapp=False)
        
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass filter values to template
        context['search_query'] = self.request.GET.get('search', '')
        context['cabang_id_filter'] = self.request.GET.get('cabang_id', '')
        context['has_whatsapp_filter'] = self.request.GET.get('has_whatsapp', '')
        # Get all cabangs for filter dropdown
        context['cabangs'] = Cabang.objects.all()
        # Get total count (with filters applied)
        context['total_pasien'] = self.get_queryset().count()
        return context

class PasienCreateView(PermissionRequiredViewMixin, CreateView):
    """Create new pasien."""
    model = Pasien
    fields = ['nama_anak', 'tanggal_lahir', 'jenis_kelamin', 'nama_orang_tua', 'alamat', 'no_wa', 'has_whatsapp', 'cabang']
    template_name = 'core/pasien_form.html'
    success_url = reverse_lazy('pasien_list')
    permission_required = 'pasien_create'
    
    def get_form(self, form_class=None):
        from django import forms
        form = super().get_form(form_class)
        form.fields['nama_anak'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Nama anak'})
        form.fields['tanggal_lahir'].widget = forms.DateInput(attrs={'class': 'form-control', 'type': 'date'})
        form.fields['jenis_kelamin'].widget = forms.Select(
            choices=[('', '-- Pilih --'), ('L', 'Laki-laki'), ('P', 'Perempuan')],
            attrs={'class': 'form-select'}
        )
        form.fields['nama_orang_tua'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Nama orang tua'})
        form.fields['alamat'].widget = forms.Textarea(attrs={'class': 'form-control', 'rows': 2, 'placeholder': 'Alamat lengkap'})
        form.fields['no_wa'].widget.attrs.update({'class': 'form-control', 'placeholder': '08xx-xxxx-xxxx', 'id': 'id_no_wa'})
        form.fields['has_whatsapp'].widget.attrs.update({'class': 'form-check-input', 'id': 'id_has_whatsapp'})
        form.fields['cabang'].widget.attrs.update({'class': 'form-select'})
        return form
    
    def form_valid(self, form):
        # Auto-generate kode_pasien
        if not form.instance.kode_pasien:
            from django.db.models import Max
            last_pasien = Pasien.objects.aggregate(Max('id'))['id__max']
            next_id = (last_pasien or 0) + 1
            form.instance.kode_pasien = f'P{next_id:04d}'
        messages.success(self.request, f'Data pasien {form.instance.nama_anak} berhasil disimpan dengan kode {form.instance.kode_pasien}!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        return super().form_invalid(form)

class PasienEditView(PermissionRequiredViewMixin, UpdateView):
    """Edit existing pasien."""
    model = Pasien
    fields = ['nama_anak', 'tanggal_lahir', 'jenis_kelamin', 'nama_orang_tua', 'alamat', 'no_wa', 'has_whatsapp', 'cabang']
    template_name = 'core/pasien_form.html'
    context_object_name = 'object'
    pk_url_kwarg = 'pk'
    permission_required = 'pasien_edit'
    
    def get_success_url(self):
        return reverse_lazy('pasien_list')
    
    def get_form(self, form_class=None):
        from django import forms
        form = super().get_form(form_class)
        form.fields['nama_anak'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Nama anak'})
        form.fields['tanggal_lahir'].widget = forms.DateInput(attrs={'class': 'form-control', 'type': 'date'})
        form.fields['jenis_kelamin'].widget = forms.Select(
            choices=[('', '-- Pilih --'), ('L', 'Laki-laki'), ('P', 'Perempuan')],
            attrs={'class': 'form-select'}
        )
        form.fields['nama_orang_tua'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Nama orang tua'})
        form.fields['alamat'].widget = forms.Textarea(attrs={'class': 'form-control', 'rows': 2, 'placeholder': 'Alamat lengkap'})
        form.fields['no_wa'].widget.attrs.update({'class': 'form-control', 'placeholder': '08xx-xxxx-xxxx', 'id': 'id_no_wa'})
        form.fields['has_whatsapp'].widget.attrs.update({'class': 'form-check-input', 'id': 'id_has_whatsapp'})
        form.fields['cabang'].widget.attrs.update({'class': 'form-select'})
        return form
    
    def form_valid(self, form):
        messages.success(self.request, f'Data pasien {form.instance.nama_anak} berhasil diperbarui!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        return super().form_invalid(form)

@require_permission('pasien_export')
def export_pasien_excel(request):
    """Export pasien list to Excel"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from datetime import datetime
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Pasien"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['No', 'Kode Pasien', 'Nama Anak', 'Tanggal Lahir', 'Usia', 'Jenis Kelamin', 
               'Nama Orang Tua', 'No. WhatsApp', 'Status WA', 'Alamat', 'Cabang']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Get data with same filters as list view
    qs = Pasien.objects.select_related('cabang').filter(is_deleted=False).order_by('-id')
    
    # Apply filters from GET parameters
    search_query = request.GET.get('search')
    if search_query:
        from django.db.models import Q
        qs = qs.filter(
            Q(nama_anak__icontains=search_query) |
            Q(nama_orang_tua__icontains=search_query) |
            Q(no_wa__icontains=search_query) |
            Q(kode_pasien__icontains=search_query)
        )
    
    cabang_id = request.GET.get('cabang_id')
    if cabang_id:
        qs = qs.filter(cabang_id=cabang_id)
    
    has_wa = request.GET.get('has_whatsapp')
    if has_wa == 'true':
        qs = qs.filter(has_whatsapp=True)
    elif has_wa == 'false':
        qs = qs.filter(has_whatsapp=False)
    
    # Fill data
    for idx, pasien in enumerate(qs, start=1):
        # Calculate age
        today = datetime.now().date()
        age = today.year - pasien.tanggal_lahir.year - ((today.month, today.day) < (pasien.tanggal_lahir.month, pasien.tanggal_lahir.day))
        
        row = [
            idx,
            pasien.kode_pasien or '-',
            pasien.nama_anak,
            pasien.tanggal_lahir.strftime('%d-%m-%Y'),
            f"{age} tahun",
            'Laki-laki' if pasien.jenis_kelamin == 'L' else 'Perempuan' if pasien.jenis_kelamin == 'P' else '-',
            pasien.nama_orang_tua or '-',
            pasien.no_wa or '-',
            '✓ Terdaftar' if pasien.has_whatsapp else '✗ Belum',
            pasien.alamat or '-',
            pasien.cabang.nama_cabang if pasien.cabang else '-'
        ]
        ws.append(row)
    
    # Adjust column widths
    column_widths = [5, 15, 20, 15, 10, 15, 20, 15, 15, 30, 15]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Data_Pasien_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


class ImportPasienView(PermissionRequiredViewMixin, View):
    """Import pasien data from Excel file."""
    permission_required = 'pasien_create'

    def get(self, request):
        """Return list of cabangs for the modal dropdown."""
        cabangs = list(Cabang.objects.values('id', 'nama_cabang').order_by('nama_cabang'))
        return JsonResponse({'cabangs': cabangs})

    def post(self, request):
        import re
        import difflib
        import openpyxl
        from django.db.models import Max

        BULAN_ID = {
            'JANUARI': 1, 'FEBRUARI': 2, 'MARET': 3, 'APRIL': 4,
            'MEI': 5, 'JUNI': 6, 'JULI': 7, 'AGUSTUS': 8,
            'SEPTEMBER': 9, 'OKTOBER': 10, 'NOVEMBER': 11, 'DESEMBER': 12,
        }

        def _resolve_bulan(raw):
            """Return month number for raw string, with fuzzy fallback (cutoff 0.7)."""
            exact = BULAN_ID.get(raw)
            if exact:
                return exact
            matches = difflib.get_close_matches(raw, BULAN_ID.keys(), n=1, cutoff=0.7)
            return BULAN_ID[matches[0]] if matches else None

        def parse_tanggal(val):
            if not val:
                return None
            val_str = str(val).strip().upper()
            match = re.search(r'(\d{1,2})\s+([A-Z]+)\s+(\d{4})', val_str)
            if match:
                try:
                    day = int(match.group(1))
                    month = _resolve_bulan(match.group(2))
                    year = int(match.group(3))
                    if month:
                        from datetime import date as date_cls
                        return date_cls(year, month, day)
                except (ValueError, TypeError):
                    pass
            return None

        def parse_jenis_kelamin(val):
            if not val:
                return None
            v = re.sub(r'[\s\-]', '', str(val).strip().upper())
            if v == 'LAKILAKI':
                return 'L'
            if v == 'PEREMPUAN':
                return 'P'
            return None

        def parse_no_wa(val):
            if not val:
                return None
            v = re.sub(r'[\s\-]', '', str(val).strip())
            return v if v else None

        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return JsonResponse({'success': False, 'error': 'File tidak ditemukan.'})

        cabang_id = request.POST.get('cabang_id') or None
        cabang = None
        if cabang_id:
            try:
                cabang = Cabang.objects.get(id=cabang_id)
            except Cabang.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Cabang tidak ditemukan.'})

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active

            # Detect header row in first 3 rows
            header_row_idx = None
            col_map = {}
            for row_idx, row in enumerate(ws.iter_rows(max_row=3, values_only=True), start=1):
                row_upper = [str(c).strip().upper() if c is not None else '' for c in row]
                if 'NAMA PASIEN' in row_upper:
                    header_row_idx = row_idx
                    for col_idx, cell_val in enumerate(row_upper):
                        col_map[cell_val] = col_idx
                    break

            if header_row_idx is None:
                return JsonResponse({
                    'success': False,
                    'error': 'Header tidak ditemukan. Pastikan baris header berisi kolom "NAMA PASIEN" dalam 3 baris pertama.'
                })

            saved = 0
            errors = []

            for row_num, row in enumerate(
                ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
                start=header_row_idx + 1
            ):
                # Skip fully empty rows
                if all(c is None or str(c).strip() == '' for c in row):
                    continue

                def get_col(name):
                    idx = col_map.get(name)
                    if idx is not None and idx < len(row):
                        v = row[idx]
                        return str(v).strip() if v is not None else ''
                    return ''

                nama_anak = get_col('NAMA PASIEN')
                if not nama_anak:
                    continue

                tanggal_lahir_val = parse_tanggal(get_col('TEMPAT TANGGAL LAHIR'))
                if tanggal_lahir_val is None:
                    errors.append(f'Baris {row_num} ("{nama_anak}"): format tanggal lahir tidak dikenali.')
                    continue

                try:
                    pasien = Pasien.objects.create(
                        nama_anak=nama_anak,
                        tanggal_lahir=tanggal_lahir_val,
                        jenis_kelamin=parse_jenis_kelamin(get_col('JENIS KELAMIN')),
                        nama_orang_tua=get_col('NAMA BUNDA') or None,
                        no_wa=parse_no_wa(get_col('NOMOR TELP')) or None,
                        cabang=cabang,
                        is_deleted=False,
                    )
                    pasien.kode_pasien = f'P{pasien.id:04d}'
                    pasien.save(update_fields=['kode_pasien'])
                    saved += 1
                except Exception as e:
                    errors.append(f'Baris {row_num} ("{nama_anak}"): {str(e)}')

            return JsonResponse({
                'success': True,
                'saved': saved,
                'errors': errors,
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Gagal membaca file: {str(e)}'})


@require_permission('registrasi_export')
def export_registrasi_excel(request):
    """Export registrasi list to Excel"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from datetime import datetime
    from django.db.models import Sum, F, DecimalField, Case, When
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Registrasi"
    
    # Header styling
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['No', 'Kode Registrasi', 'Tanggal', 'Pasien', 'Jenis Terapi', 'Terapis', 
               'Total Bayar', 'Sudah Dibayar', 'Sisa Tagihan', 'Status', 'Cabang']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Get data with same filters as list view
    qs = Registrasi.objects.select_related(
        'pasien', 'jenis_terapi', 'terapis', 'cabang'
    ).order_by('-tanggal_kunjungan')
    
    # Add payment info annotations
    qs = qs.annotate(
        total_paid=Sum('pemasukan__jumlah', output_field=DecimalField())
    )
    qs = qs.annotate(
        sisa_tagihan=Case(
            When(total_paid__isnull=True, then=F('total_bayar')),
            default=F('total_bayar') - F('total_paid'),
            output_field=DecimalField()
        )
    )
    
    # Apply filters from GET parameters
    tanggal_dari = request.GET.get('tanggal_dari')
    tanggal_sampai = request.GET.get('tanggal_sampai')
    
    if tanggal_dari:
        try:
            tanggal_dari_date = datetime.strptime(tanggal_dari, '%Y-%m-%d').date()
            qs = qs.filter(tanggal_kunjungan__gte=tanggal_dari_date)
        except:
            pass
    
    if tanggal_sampai:
        try:
            tanggal_sampai_date = datetime.strptime(tanggal_sampai, '%Y-%m-%d').date()
            qs = qs.filter(tanggal_kunjungan__lte=tanggal_sampai_date)
        except:
            pass
    
    pasien_query = request.GET.get('pasien_query')
    if pasien_query:
        qs = qs.filter(pasien__nama_anak__icontains=pasien_query)
    
    jenis_terapi_id = request.GET.get('jenis_terapi_id')
    if jenis_terapi_id:
        qs = qs.filter(jenis_terapi_id=jenis_terapi_id)
    
    terapis_id = request.GET.get('terapis_id')
    if terapis_id:
        qs = qs.filter(terapis_id=terapis_id)
    
    # Fill data
    for idx, reg in enumerate(qs, start=1):
        total_paid = reg.total_paid or 0
        sisa = reg.sisa_tagihan or 0
        status = 'LUNAS' if sisa <= 0 else 'BELUM LUNAS'
        
        row = [
            idx,
            reg.kode_registrasi or '-',
            reg.tanggal_kunjungan.strftime('%d-%m-%Y'),
            reg.pasien.nama_anak,
            reg.jenis_terapi.nama_terapi,
            reg.terapis.nama_terapis if reg.terapis else '-',
            float(reg.total_bayar),
            float(total_paid),
            float(sisa),
            status,
            reg.cabang.nama_cabang if reg.cabang else '-'
        ]
        ws.append(row)
    
    # Format currency columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=9):
        for cell in row:
            cell.number_format = '"Rp "#,##0'
    
    # Adjust column widths
    column_widths = [5, 18, 12, 20, 25, 20, 15, 15, 15, 15, 15]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Data_Registrasi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

class TerapisListView(LoginRequiredMixin, ListView):
    """List all terapis."""
    model = Terapis
    template_name = 'core/terapis_list.html'
    context_object_name = 'terapis_list'
    paginate_by = 25
    
    def get_queryset(self):
        qs = super().get_queryset().select_related('cabang').order_by('-id')
        if getattr(self.request, 'cabang_id', None) is not None:
            qs = qs.filter(cabang_id=self.request.cabang_id)
        
        # Search by nama_terapis or no_hp
        search_query = self.request.GET.get('search')
        if search_query:
            from django.db.models import Q
            qs = qs.filter(
                Q(nama_terapis__icontains=search_query) |
                Q(no_hp__icontains=search_query)
            )
        
        # Filter by cabang
        cabang_id = self.request.GET.get('cabang_id')
        if cabang_id:
            qs = qs.filter(cabang_id=cabang_id)
        
        # Filter by active status
        is_active = self.request.GET.get('is_active')
        if is_active == 'true':
            qs = qs.filter(is_active=True, is_deleted=False)
        elif is_active == 'false':
            qs = qs.filter(is_active=False)
        
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass filter values to template
        context['search_query'] = self.request.GET.get('search', '')
        context['cabang_id_filter'] = self.request.GET.get('cabang_id', '')
        context['is_active_filter'] = self.request.GET.get('is_active', '')
        # Get all cabangs for filter dropdown
        context['cabangs'] = Cabang.objects.all()
        # Get total count and active count
        context['total_terapis'] = self.get_queryset().count()
        context['total_aktif'] = self.get_queryset().filter(is_active=True, is_deleted=False).count()
        return context

class TerapisCreateView(LoginRequiredMixin, CreateView):
    """Create new terapis."""
    model = Terapis
    fields = ['nama_terapis', 'no_hp', 'alamat', 'cabang', 'biaya_transport_default', 'is_active']
    template_name = 'core/terapis_form.html'
    success_url = reverse_lazy('terapis_list')
    
    def get_form(self, form_class=None):
        from django import forms
        form = super().get_form(form_class)
        form.fields['nama_terapis'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Nama terapis'})
        form.fields['no_hp'].widget.attrs.update({'class': 'form-control', 'placeholder': '08xx-xxxx-xxxx'})
        form.fields['alamat'].widget = forms.Textarea(attrs={'class': 'form-control', 'rows': 2, 'placeholder': 'Alamat lengkap'})
        form.fields['cabang'].widget.attrs.update({'class': 'form-select'})
        form.fields['biaya_transport_default'].widget.attrs.update({'class': 'form-control', 'placeholder': '0'})
        form.fields['is_active'].widget.attrs.update({'class': 'form-check-input', 'checked': 'checked'})
        form.fields['is_active'].initial = True
        form.fields['biaya_transport_default'].initial = 0
        return form
    
    def form_valid(self, form):
        messages.success(self.request, f'Data terapis {form.instance.nama_terapis} berhasil disimpan!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        return super().form_invalid(form)

class TerapisUpdateView(LoginRequiredMixin, UpdateView):
    """Update terapis."""
    model = Terapis
    fields = ['nama_terapis', 'no_hp', 'alamat', 'cabang', 'biaya_transport_default', 'is_active']
    template_name = 'core/terapis_form.html'
    success_url = reverse_lazy('terapis_list')
    
    def get_form(self, form_class=None):
        from django import forms
        form = super().get_form(form_class)
        form.fields['nama_terapis'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Nama terapis'})
        form.fields['no_hp'].widget.attrs.update({'class': 'form-control', 'placeholder': '08xx-xxxx-xxxx'})
        form.fields['alamat'].widget = forms.Textarea(attrs={'class': 'form-control', 'rows': 2, 'placeholder': 'Alamat lengkap'})
        form.fields['cabang'].widget.attrs.update({'class': 'form-select'})
        form.fields['biaya_transport_default'].widget.attrs.update({'class': 'form-control', 'placeholder': '0'})
        form.fields['is_active'].widget.attrs.update({'class': 'form-check-input'})
        return form
    
    def form_valid(self, form):
        messages.success(self.request, f'Data terapis {form.instance.nama_terapis} berhasil diperbarui!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        return super().form_invalid(form)

class TerapisDeleteView(LoginRequiredMixin, DeleteView):
    """Delete terapis."""
    model = Terapis
    success_url = reverse_lazy('terapis_list')
    
    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        nama = obj.nama_terapis
        messages.success(request, f'Data terapis {nama} berhasil dihapus!')
        return super().delete(request, *args, **kwargs)

class JenisTerapiListView(LoginRequiredMixin, ListView):
    """List all jenis terapi."""
    model = JenisTerapi
    template_name = 'core/jenis_terapi_list.html'
    context_object_name = 'jenis_terapi_list'
    paginate_by = 25

class JenisTerapiCreateView(LoginRequiredMixin, CreateView):
    """Create new jenis terapi."""
    model = JenisTerapi
    fields = ['nama_terapi', 'kategori_usia_min', 'kategori_usia_max', 'harga']
    template_name = 'core/jenis_terapi_form.html'
    success_url = reverse_lazy('jenis_terapi_list')
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields['nama_terapi'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Nama jenis terapi'})
        form.fields['kategori_usia_min'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Min. umur (bulan)'})
        form.fields['kategori_usia_max'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Max. umur (bulan)'})
        form.fields['harga'].widget.attrs.update({'class': 'form-control', 'placeholder': '0'})
        return form
    
    def form_valid(self, form):
        messages.success(self.request, f'Jenis terapi {form.instance.nama_terapi} berhasil disimpan!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        return super().form_invalid(form)

class JenisTerapiUpdateView(LoginRequiredMixin, UpdateView):
    """Update jenis terapi."""
    model = JenisTerapi
    fields = ['nama_terapi', 'kategori_usia_min', 'kategori_usia_max', 'harga']
    template_name = 'core/jenis_terapi_form.html'
    success_url = reverse_lazy('jenis_terapi_list')
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields['nama_terapi'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Nama jenis terapi'})
        form.fields['kategori_usia_min'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Min. umur (bulan)'})
        form.fields['kategori_usia_max'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Max. umur (bulan)'})
        form.fields['harga'].widget.attrs.update({'class': 'form-control', 'placeholder': '0'})
        return form
    
    def form_valid(self, form):
        messages.success(self.request, f'Jenis terapi {form.instance.nama_terapi} berhasil diperbarui!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        return super().form_invalid(form)

class JenisTerapiDeleteView(LoginRequiredMixin, DeleteView):
    """Delete jenis terapi."""
    model = JenisTerapi
    success_url = reverse_lazy('jenis_terapi_list')
    
    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        nama = obj.nama_terapi
        messages.success(request, f'Jenis terapi {nama} berhasil dihapus!')
        return super().delete(request, *args, **kwargs)

class CabangListView(LoginRequiredMixin, ListView):
    """List all cabang."""
    model = Cabang
    template_name = 'core/cabang_list.html'
    context_object_name = 'cabang_list'
    paginate_by = 25

class CabangCreateView(LoginRequiredMixin, CreateView):
    """Create new cabang."""
    model = Cabang
    fields = ['nama_cabang', 'alamat']
    template_name = 'core/cabang_form.html'
    success_url = reverse_lazy('cabang_list')
    
    def get_form(self, form_class=None):
        from django import forms
        form = super().get_form(form_class)
        form.fields['nama_cabang'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Nama cabang'})
        form.fields['alamat'].widget = forms.Textarea(attrs={'class': 'form-control', 'rows': 2, 'placeholder': 'Alamat lengkap'})
        return form
    
    def form_valid(self, form):
        messages.success(self.request, f'Cabang {form.instance.nama_cabang} berhasil disimpan!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        return super().form_invalid(form)

class CabangUpdateView(LoginRequiredMixin, UpdateView):
    """Update existing cabang."""
    model = Cabang
    fields = ['nama_cabang', 'alamat']
    template_name = 'core/cabang_form.html'
    success_url = reverse_lazy('cabang_list')
    
    def get_form(self, form_class=None):
        from django import forms
        form = super().get_form(form_class)
        form.fields['nama_cabang'].widget.attrs.update({'class': 'form-control', 'placeholder': 'Nama cabang'})
        form.fields['alamat'].widget = forms.Textarea(attrs={'class': 'form-control', 'rows': 2, 'placeholder': 'Alamat lengkap'})
        return form
    
    def form_valid(self, form):
        messages.success(self.request, f'Cabang {form.instance.nama_cabang} berhasil diperbarui!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Terjadi kesalahan. Periksa kembali data yang diinput.')
        return super().form_invalid(form)

class CabangDeleteView(LoginRequiredMixin, DeleteView):
    """Delete existing cabang."""
    model = Cabang
    success_url = reverse_lazy('cabang_list')
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        nama_cabang = self.object.nama_cabang
        response = super().delete(request, *args, **kwargs)
        messages.success(self.request, f'Cabang {nama_cabang} berhasil dihapus!')
        return response


class RekapTindakanListView(PermissionRequiredViewMixin, ListView):
    """Rekap tindakan with filtering"""
    model = Registrasi
    template_name = 'core/rekap_tindakan_list.html'
    context_object_name = 'rekap_list'
    paginate_by = 50
    permission_required = 'rekap_view'

    def get_queryset(self):
        today = timezone.now().date()
        qs = Registrasi.objects.select_related(
            'pasien', 'jenis_terapi', 'terapis', 'cabang'
        ).order_by('-tanggal_kunjungan')
        
        period = self.request.GET.get('period', 'harian')
        
        if period == 'harian':
            qs = qs.filter(tanggal_kunjungan=today)
        elif period == 'minggu':
            start_date = today - timedelta(days=6)
            qs = qs.filter(tanggal_kunjungan__gte=start_date, tanggal_kunjungan__lte=today)
        elif period == '3bulan':
            start_date = today - timedelta(days=89)
            qs = qs.filter(tanggal_kunjungan__gte=start_date, tanggal_kunjungan__lte=today)
        elif period == '6bulan':
            start_date = today - timedelta(days=179)
            qs = qs.filter(tanggal_kunjungan__gte=start_date, tanggal_kunjungan__lte=today)
        elif period == '1tahun':
            start_date = today - timedelta(days=364)
            qs = qs.filter(tanggal_kunjungan__gte=start_date, tanggal_kunjungan__lte=today)
        
        # Custom date range if provided
        start_custom = self.request.GET.get('start_date')
        end_custom = self.request.GET.get('end_date')
        if start_custom and end_custom:
            try:
                start_date = datetime.strptime(start_custom, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_custom, '%Y-%m-%d').date()
                qs = qs.filter(tanggal_kunjungan__gte=start_date, tanggal_kunjungan__lte=end_date)
            except ValueError:
                pass
        
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get same queryset for totals (all records, not just current page)
        qs = self.get_queryset()
        
        # Calculate totals
        total_harga = qs.aggregate(total=Sum('harga'))['total'] or 0
        total_transport = qs.aggregate(total=Sum('biaya_transport'))['total'] or 0
        total_pendapatan = qs.aggregate(total=Sum('total_bayar'))['total'] or 0
        
        context['total_harga'] = total_harga
        context['total_transport'] = total_transport
        context['total_pendapatan'] = total_pendapatan
        context['current_period'] = self.request.GET.get('period', 'harian')
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        context['record_count'] = qs.count()
        
        # Attach payment method to each item in current page
        current_page_items = context['rekap_list']
        registrasi_ids = [item.id for item in current_page_items]
        if registrasi_ids:
            pemasukan_records = Pemasukan.objects.filter(
                registrasi_id__in=registrasi_ids
            ).select_related('registrasi')
            pemasukan_dict = {p.registrasi_id: p.metode_pembayaran for p in pemasukan_records}
            
            # Add payment method as attribute to each item
            for item in current_page_items:
                item.metode_pembayaran = pemasukan_dict.get(item.id, None)
        
        return context


def build_whatsapp_url(phone_number, message_text):
    formatted_phone = whatsapp_service.format_phone_number(phone_number)
    if not formatted_phone:
        return None
    encoded_message = quote(message_text or '')
    return f'https://wa.me/{formatted_phone}?text={encoded_message}'


class NotifikasiListView(PermissionRequiredViewMixin, ListView):
    model = Notifikasi
    template_name = 'core/notifikasi_list.html'
    context_object_name = 'notifikasi_list'
    paginate_by = 25
    permission_required = 'notifikasi_view'
    
    def get_queryset(self):
        qs = Notifikasi.objects.select_related('pasien', 'registrasi').order_by('-tanggal_notifikasi', '-created_at')
        
        # Filter by read/unread
        filter_type = self.request.GET.get('filter', 'all')
        if filter_type == 'unread':
            qs = qs.filter(sudah_dibaca=False)
        elif filter_type == 'read':
            qs = qs.filter(sudah_dibaca=True)
        
        # Filter by type
        jenis = self.request.GET.get('jenis')
        if jenis:
            qs = qs.filter(jenis_notifikasi=jenis)
        
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unread_count'] = Notifikasi.objects.filter(sudah_dibaca=False).count()
        context['current_filter'] = self.request.GET.get('filter', 'all')
        context['current_jenis'] = self.request.GET.get('jenis', '')
        context['jenis_list'] = NOTIFICATION_TYPE_CHOICES
        context['can_manage_templates'] = self.request.user.has_permission('template_pesan_view')
        context['can_generate_notifications'] = self.request.user.has_permission('notifikasi_generate')
        context['can_mark_notifications'] = self.request.user.has_permission('notifikasi_mark_read')

        type_label_map = dict(NOTIFICATION_TYPE_CHOICES)
        for notif in context['notifikasi_list']:
            notif.jenis_label = type_label_map.get(notif.jenis_notifikasi, notif.jenis_notifikasi or 'Info')
            notif.whatsapp_number = getattr(notif.pasien, 'no_wa', None)
            notif.rendered_message = TemplatePesan.build_message_for_notification(notif)
            notif.whatsapp_url = None
            if notif.whatsapp_number and notif.rendered_message:
                notif.whatsapp_url = build_whatsapp_url(notif.whatsapp_number, notif.rendered_message)
        
        return context


class MarkNotifikasiReadView(PermissionRequiredViewMixin, View):
    """Mark notification as read via AJAX"""
    permission_required = 'notifikasi_mark_read'

    def post(self, request, pk):
        try:
            notifikasi = Notifikasi.objects.get(pk=pk)
            notifikasi.sudah_dibaca = True
            notifikasi.save()
            return JsonResponse({'success': True})
        except Notifikasi.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Notifikasi tidak ditemukan'}, status=404)


class MarkAllNotifikasiReadView(PermissionRequiredViewMixin, View):
    """Mark all notifications as read"""
    permission_required = 'notifikasi_mark_read'

    def post(self, request):
        Notifikasi.objects.filter(sudah_dibaca=False).update(sudah_dibaca=True)
        messages.success(request, 'Semua notifikasi telah ditandai sebagai dibaca')
        return redirect('notifikasi_list')


class GenerateNotificationsManualView(PermissionRequiredViewMixin, View):
    """Manually trigger notification generation (birthday, inactive patients, etc)"""
    permission_required = 'notifikasi_generate'

    def post(self, request):
        try:
            result = generate_all_notifications()
            total_created = result.get('total_created', 0)
            
            if total_created > 0:
                messages.success(
                    request, 
                    f'✓ Berhasil membuat {total_created} notifikasi baru! '
                    f'Cek detail di halaman notifikasi.'
                )
            else:
                messages.info(
                    request, 
                    'Tidak ada notifikasi baru yang perlu dibuat saat ini.'
                )
        except Exception as e:
            messages.error(request, f'✗ Gagal generate notifikasi: {str(e)}')
        
        return redirect('notifikasi_list')


class TemplatePesanListView(PermissionRequiredViewMixin, ListView):
    model = TemplatePesan
    template_name = 'core/template_pesan_list.html'
    context_object_name = 'template_list'
    permission_required = 'template_pesan_view'

    def get_queryset(self):
        type_order = {code: index for index, (code, _label) in enumerate(NOTIFICATION_TYPE_CHOICES)}
        type_labels = dict(NOTIFICATION_TYPE_CHOICES)
        templates = list(TemplatePesan.objects.all())
        for item in templates:
            item.label = type_labels.get(item.tipe_pesan, item.tipe_pesan)
        templates.sort(key=lambda item: type_order.get(item.tipe_pesan, 999))
        return templates

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['can_edit_templates'] = self.request.user.has_permission('template_pesan_edit')
        return context


class TemplatePesanCreateView(PermissionRequiredViewMixin, CreateView):
    model = TemplatePesan
    form_class = TemplatePesanForm
    template_name = 'core/template_pesan_form.html'
    success_url = reverse_lazy('template_pesan_list')
    permission_required = 'template_pesan_edit'

    def form_valid(self, form):
        messages.success(self.request, 'Template pesan berhasil dibuat.')
        return super().form_valid(form)


class TemplatePesanEditView(PermissionRequiredViewMixin, UpdateView):
    model = TemplatePesan
    form_class = TemplatePesanForm
    template_name = 'core/template_pesan_form.html'
    success_url = reverse_lazy('template_pesan_list')
    permission_required = 'template_pesan_edit'

    def form_valid(self, form):
        messages.success(self.request, 'Template pesan berhasil diperbarui.')
        return super().form_valid(form)


class TemplatePesanDeleteView(PermissionRequiredViewMixin, View):
    permission_required = 'template_pesan_edit'

    def post(self, request, pk):
        template = get_object_or_404(TemplatePesan, pk=pk)
        template.delete()
        messages.success(request, 'Template pesan berhasil dihapus.')
        return redirect('template_pesan_list')


# AJAX Views for Quick Create from Modals
class AjaxCreatePasienView(LoginRequiredMixin, View):
    """Create pasien via AJAX from modal"""
    def post(self, request):
        try:
            # Get form data
            nama_anak = request.POST.get('nama_anak', '').strip()
            tanggal_lahir = request.POST.get('tanggal_lahir', '').strip()
            jenis_kelamin = request.POST.get('jenis_kelamin', '').strip()
            nama_orang_tua = request.POST.get('nama_orang_tua', '').strip()
            alamat = request.POST.get('alamat', '').strip()
            no_wa = request.POST.get('no_wa', '').strip()
            cabang_id = request.POST.get('cabang', '').strip()
            
            # Validate required fields
            if not nama_anak or not tanggal_lahir or not jenis_kelamin:
                return JsonResponse({
                    'success': False, 
                    'error': 'Nama anak, tanggal lahir, dan jenis kelamin harus diisi'
                }, status=400)
            
            # Auto-generate kode_pasien
            from django.db.models import Max
            last_pasien = Pasien.objects.aggregate(Max('id'))['id__max']
            next_id = (last_pasien or 0) + 1
            kode_pasien = f'P{next_id:04d}'
            
            # Create pasien
            pasien = Pasien.objects.create(
                kode_pasien=kode_pasien,
                nama_anak=nama_anak,
                tanggal_lahir=tanggal_lahir,
                jenis_kelamin=jenis_kelamin,
                nama_orang_tua=nama_orang_tua if nama_orang_tua else None,
                alamat=alamat if alamat else None,
                no_wa=no_wa if no_wa else None,
                cabang_id=int(cabang_id) if cabang_id else None
            )
            
            return JsonResponse({
                'success': True,
                'pasien': {
                    'id': pasien.id,
                    'text': f'{pasien.kode_pasien} - {pasien.nama_anak}'
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class AjaxCreateTerapisView(LoginRequiredMixin, View):
    """Create terapis via AJAX from modal"""
    def post(self, request):
        try:
            # Get form data
            nama_terapis = request.POST.get('nama_terapis', '').strip()
            no_hp = request.POST.get('no_hp', '').strip()
            alamat = request.POST.get('alamat', '').strip()
            cabang_id = request.POST.get('cabang', '').strip()
            biaya_transport_default = request.POST.get('biaya_transport_default', '0').strip()
            is_active = request.POST.get('is_active') == 'on'
            
            # Validate required fields
            if not nama_terapis:
                return JsonResponse({
                    'success': False,
                    'error': 'Nama terapis harus diisi'
                }, status=400)
            
            # Create terapis
            terapis = Terapis.objects.create(
                nama_terapis=nama_terapis,
                no_hp=no_hp if no_hp else None,
                alamat=alamat if alamat else None,
                cabang_id=int(cabang_id) if cabang_id else None,
                biaya_transport_default=int(biaya_transport_default.replace('.', '')) if biaya_transport_default else 0,
                is_active=is_active
            )
            
            return JsonResponse({
                'success': True,
                'terapis': {
                    'id': terapis.id,
                    'text': terapis.nama_terapis
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class AjaxCreateJenisTerapiView(LoginRequiredMixin, View):
    """Create jenis terapi via AJAX from modal"""
    def post(self, request):
        try:
            # Get form data
            nama_terapi = request.POST.get('nama_terapi', '').strip()
            kategori_usia_min = request.POST.get('kategori_usia_min', '').strip()
            kategori_usia_max = request.POST.get('kategori_usia_max', '').strip()
            harga = request.POST.get('harga', '0').strip()
            
            # Validate required fields
            if not nama_terapi:
                return JsonResponse({
                    'success': False,
                    'error': 'Nama terapi harus diisi'
                }, status=400)
            
            # Create jenis terapi
            jenis_terapi = JenisTerapi.objects.create(
                nama_terapi=nama_terapi,
                kategori_usia_min=int(kategori_usia_min) if kategori_usia_min else None,
                kategori_usia_max=int(kategori_usia_max) if kategori_usia_max else None,
                harga=int(harga.replace('.', '')) if harga else 0
            )
            
            return JsonResponse({
                'success': True,
                'jenis_terapi': {
                    'id': jenis_terapi.id,
                    'text': jenis_terapi.nama_terapi
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class AjaxGetJenisTerapiPriceView(LoginRequiredMixin, View):
    """Get jenis terapi price via AJAX"""
    def get(self, request, jenis_terapi_id):
        try:
            jenis_terapi = JenisTerapi.objects.get(pk=jenis_terapi_id)
            return JsonResponse({
                'success': True,
                'harga': int(jenis_terapi.harga)  # Convert Decimal to int
            })
        except JenisTerapi.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Jenis terapi tidak ditemukan'
            }, status=404)

class AjaxGetTerapisTransportView(LoginRequiredMixin, View):
    """Get terapis default transport cost via AJAX"""
    def get(self, request, terapis_id):
        try:
            terapis = Terapis.objects.get(pk=terapis_id)
            return JsonResponse({
                'success': True,
                'biaya_transport': int(terapis.biaya_transport_default or 0)  # Convert Decimal to int
            })
        except Terapis.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Terapis tidak ditemukan'
            }, status=404)


class AjaxGetRegistrasiDetailView(LoginRequiredMixin, View):
    """Get registrasi detail including payment info via AJAX"""
    def get(self, request, registrasi_id):
        try:
            from django.db.models import Sum
            registrasi = Registrasi.objects.select_related('pasien', 'jenis_terapi', 'terapis', 'cabang').get(pk=registrasi_id)
            
            # Calculate total already paid
            total_paid = Pemasukan.objects.filter(registrasi_id=registrasi_id).aggregate(
                total=Sum('jumlah')
            )['total'] or 0
            
            total_bayar = float(registrasi.total_bayar or 0)
            sisa = total_bayar - float(total_paid)
            
            return JsonResponse({
                'success': True,
                'registrasi': {
                    'id': registrasi.id,
                    'kode': registrasi.kode_registrasi,
                    'pasien': registrasi.pasien.nama_anak,
                    'jenis_terapi': registrasi.jenis_terapi.nama_terapi,
                    'terapis': registrasi.terapis.nama_terapis if registrasi.terapis else '-',
                    'tanggal_kunjungan': registrasi.tanggal_kunjungan.strftime('%d %b %Y'),
                    'cabang_name': registrasi.cabang.nama_cabang,
                    'total_bayar': total_bayar,
                    'total_paid': float(total_paid),
                    'sisa': sisa,
                }
            })
        except Registrasi.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Registrasi tidak ditemukan'
            }, status=404)


# ============================================================================
# PEMBUKUAN (Accounting) Views
# ============================================================================

class TotalPendapatanView(PermissionRequiredViewMixin, TemplateView):
    """Total income report"""
    template_name = 'core/pembukuan/total_pendapatan.html'
    permission_required = 'pembukuan_view'
    
    def get(self, request, *args, **kwargs):
        """Handle both normal view and Excel export"""
        if request.GET.get('export') == 'excel':
            return self.export_to_excel(request)
        return super().get(request, *args, **kwargs)
    
    def apply_filters(self, queryset, request):
        """Apply filters from GET parameters"""
        from django.db.models import Q
        from datetime import datetime
        
        # Date range filter
        tanggal_dari = request.GET.get('tanggal_dari')
        tanggal_sampai = request.GET.get('tanggal_sampai')
        
        if tanggal_dari:
            try:
                tanggal_dari_date = datetime.strptime(tanggal_dari, '%Y-%m-%d').date()
                queryset = queryset.filter(tanggal__gte=tanggal_dari_date)
            except:
                pass
        
        if tanggal_sampai:
            try:
                tanggal_sampai_date = datetime.strptime(tanggal_sampai, '%Y-%m-%d').date()
                queryset = queryset.filter(tanggal__lte=tanggal_sampai_date)
            except:
                pass
        
        # Cabang filter
        cabang_id = request.GET.get('cabang_id')
        if cabang_id:
            queryset = queryset.filter(cabang_id=cabang_id)
        
        # Metode Pembayaran filter
        metode = request.GET.get('metode_pembayaran')
        if metode:
            queryset = queryset.filter(metode_pembayaran=metode)
        
        return queryset
    
    def export_to_excel(self, request):
        """Export pemasukan data to Excel file"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        
        # Get data with proper joins and apply filters
        pemasukan_list = Pemasukan.objects.select_related(
            'registrasi__pasien',
            'cabang'
        ).all().order_by('-tanggal')
        
        # Apply filters
        pemasukan_list = self.apply_filters(pemasukan_list, request)
        for pemasukan in pemasukan_list:
            if pemasukan.jumlah_bayar and pemasukan.jumlah and pemasukan.jumlah_bayar > pemasukan.jumlah:
                pemasukan.kembalian = pemasukan.jumlah_bayar - pemasukan.jumlah
            else:
                pemasukan.kembalian = None
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Pemasukan"
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        
        # Style definitions
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="2c3e50")
        border = Border(
            left=Side(style='thin', color='e9ecef'),
            right=Side(style='thin', color='e9ecef'),
            top=Side(style='thin', color='e9ecef'),
            bottom=Side(style='thin', color='e9ecef')
        )
        
        # Title
        ws['A1'] = "LAPORAN PENDAPATAN"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Date
        ws['A2'] = f"Periode: {datetime.now().strftime('%d %B %Y')}"
        ws.merge_cells('A2:G2')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Tanggal', 'Pasien', 'Metode', 'Jumlah Tagihan', 'Jumlah Bayar', 'Kembalian', 'Cabang']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[4].height = 20
        
        # Data rows
        for row, pemasukan in enumerate(pemasukan_list, start=5):
            # Get pasien name via join
            pasien_name = '-'
            if pemasukan.registrasi and pemasukan.registrasi.pasien:
                pasien_name = pemasukan.registrasi.pasien.nama_anak
            
            ws.cell(row=row, column=1).value = pemasukan.tanggal.strftime('%d/%m/%Y') if pemasukan.tanggal else '-'
            ws.cell(row=row, column=2).value = pasien_name
            ws.cell(row=row, column=3).value = pemasukan.metode_pembayaran or '-'
            ws.cell(row=row, column=4).value = float(pemasukan.jumlah) if pemasukan.jumlah else 0
            ws.cell(row=row, column=5).value = float(pemasukan.jumlah_bayar) if pemasukan.jumlah_bayar else 0
            ws.cell(row=row, column=6).value = float(pemasukan.kembalian) if pemasukan.kembalian else 0
            ws.cell(row=row, column=7).value = pemasukan.cabang.nama_cabang if pemasukan.cabang else '-'
            
            # Format currency columns
            for col in [4, 5, 6]:
                ws.cell(row=row, column=col).number_format = '#,##0'
            
            # Apply border to all cells
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='left' if col <= 3 else 'right', vertical='center')
        
        # Response
        from django.http import HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Laporan_Pendapatan_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get all pemasukan with select_related for optimization
        pemasukan_list = Pemasukan.objects.select_related(
            'registrasi__pasien',
            'cabang'
        ).all().order_by('-tanggal')
        
        # Apply filters from GET parameters
        pemasukan_list = self.apply_filters(pemasukan_list, self.request)
        
        # Add computed kembalian field to each pemasukan
        for pemasukan in pemasukan_list:
            if pemasukan.jumlah_bayar and pemasukan.jumlah and pemasukan.jumlah_bayar > pemasukan.jumlah:
                pemasukan.kembalian = pemasukan.jumlah_bayar - pemasukan.jumlah
                pemasukan.kembalian_formatted = format_rupiah(int(pemasukan.kembalian))
            else:
                pemasukan.kembalian = None
                pemasukan.kembalian_formatted = None
        
        total = pemasukan_list.aggregate(Sum('jumlah'))['jumlah__sum'] or 0
        count = pemasukan_list.count()
        avg = int(total / count) if count > 0 else 0
        
        # Get filter values from request
        tanggal_dari = self.request.GET.get('tanggal_dari', '')
        tanggal_sampai = self.request.GET.get('tanggal_sampai', '')
        cabang_id = self.request.GET.get('cabang_id', '')
        metode_pembayaran = self.request.GET.get('metode_pembayaran', '')
        
        # Get dinamis cabang list
        from core.models import Cabang
        cabang_list = Cabang.objects.all()
        
        # Metode pembayaran choices
        metode_choices = [
            ('TUNAI', 'Tunai'),
            ('TRANSFER', 'Transfer'),
            ('QRIS', 'QRIS'),
            ('DEBIT', 'Debit'),
            ('KREDIT', 'Kredit'),
        ]
        
        context['pemasukan_list'] = pemasukan_list
        context['total_pendapatan'] = total
        context['total_pendapatan_formatted'] = format_rupiah(total)
        context['avg_pendapatan'] = avg
        context['avg_pendapatan_formatted'] = format_rupiah(avg)
        context['cabang_list'] = cabang_list
        context['metode_choices'] = metode_choices
        context['tanggal_dari'] = tanggal_dari
        context['tanggal_sampai'] = tanggal_sampai
        context['cabang_id'] = cabang_id
        context['metode_pembayaran'] = metode_pembayaran
        return context


class TotalPengeluaranView(PermissionRequiredViewMixin, TemplateView):
    """Total expense report"""
    template_name = 'core/pembukuan/total_pengeluaran.html'
    permission_required = 'pembukuan_view'
    
    def get(self, request, *args, **kwargs):
        """Handle both normal view and Excel export"""
        if request.GET.get('export') == 'excel':
            return self.export_to_excel(request)
        return super().get(request, *args, **kwargs)
    
    def apply_filters(self, queryset, request):
        """Apply filters from GET parameters"""
        from django.db.models import Q
        from datetime import datetime
        
        # Date range filter
        tanggal_dari = request.GET.get('tanggal_dari')
        tanggal_sampai = request.GET.get('tanggal_sampai')
        
        if tanggal_dari:
            try:
                tanggal_dari_date = datetime.strptime(tanggal_dari, '%Y-%m-%d').date()
                queryset = queryset.filter(tanggal__gte=tanggal_dari_date)
            except:
                pass
        
        if tanggal_sampai:
            try:
                tanggal_sampai_date = datetime.strptime(tanggal_sampai, '%Y-%m-%d').date()
                queryset = queryset.filter(tanggal__lte=tanggal_sampai_date)
            except:
                pass
        
        # Cabang filter
        cabang_id = request.GET.get('cabang_id')
        if cabang_id:
            queryset = queryset.filter(cabang_id=cabang_id)
        
        # Kategori filter
        kategori = request.GET.get('kategori')
        if kategori:
            queryset = queryset.filter(kategori=kategori)
        
        return queryset
    
    def export_to_excel(self, request):
        """Export pengeluaran data to Excel file"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        
        # Get data with proper joins and apply filters
        pengeluaran_list = Pengeluaran.objects.select_related('cabang').all().order_by('-tanggal')
        
        # Apply filters
        pengeluaran_list = self.apply_filters(pengeluaran_list, request)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Pengeluaran"
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        
        # Style definitions
        header_fill = PatternFill(start_color="f5576c", end_color="f5576c", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="2c3e50")
        border = Border(
            left=Side(style='thin', color='e9ecef'),
            right=Side(style='thin', color='e9ecef'),
            top=Side(style='thin', color='e9ecef'),
            bottom=Side(style='thin', color='e9ecef')
        )
        
        # Title
        ws['A1'] = "LAPORAN PENGELUARAN"
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Date
        ws['A2'] = f"Periode: {datetime.now().strftime('%d %B %Y')}"
        ws.merge_cells('A2:E2')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Tanggal', 'Kategori', 'Keterangan', 'Cabang', 'Jumlah']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[4].height = 20
        
        # Data rows
        for row, pengeluaran in enumerate(pengeluaran_list, start=5):
            ws.cell(row=row, column=1).value = pengeluaran.tanggal.strftime('%d/%m/%Y') if pengeluaran.tanggal else '-'
            ws.cell(row=row, column=2).value = pengeluaran.kategori or '-'
            ws.cell(row=row, column=3).value = pengeluaran.keterangan or ''
            ws.cell(row=row, column=4).value = pengeluaran.cabang.nama_cabang if pengeluaran.cabang else '-'
            ws.cell(row=row, column=5).value = float(pengeluaran.jumlah) if pengeluaran.jumlah else 0
            
            # Format currency column
            ws.cell(row=row, column=5).number_format = '#,##0'
            
            # Apply border to all cells
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='left' if col <= 4 else 'right', vertical='center')
        
        # Response
        from django.http import HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Laporan_Pengeluaran_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get all pengeluaran with select_related for optimization
        pengeluaran_list = Pengeluaran.objects.select_related('cabang').all().order_by('-tanggal')
        
        # Apply filters from GET parameters
        pengeluaran_list = self.apply_filters(pengeluaran_list, self.request)
        
        total = pengeluaran_list.aggregate(Sum('jumlah'))['jumlah__sum'] or 0
        count = pengeluaran_list.count()
        avg = int(total / count) if count > 0 else 0
        
        # Get filter values from request
        tanggal_dari = self.request.GET.get('tanggal_dari', '')
        tanggal_sampai = self.request.GET.get('tanggal_sampai', '')
        cabang_id = self.request.GET.get('cabang_id', '')
        kategori_filter = self.request.GET.get('kategori', '')
        
        # Get dinamis cabang list
        from core.models import Cabang
        cabang_list = Cabang.objects.all()
        
        # Get unique kategori from pengeluaran
        kategori_choices = Pengeluaran.objects.filter(kategori__isnull=False).values_list('kategori', flat=True).distinct().order_by('kategori')
        
        context['pengeluaran_list'] = pengeluaran_list
        context['total_pengeluaran'] = total
        context['total_pengeluaran_formatted'] = format_rupiah(total)
        context['avg_pengeluaran'] = avg
        context['avg_pengeluaran_formatted'] = format_rupiah(avg)
        context['cabang_list'] = cabang_list
        context['kategori_choices'] = kategori_choices
        context['tanggal_dari'] = tanggal_dari
        context['tanggal_sampai'] = tanggal_sampai
        context['cabang_id'] = cabang_id
        context['kategori'] = kategori_filter
        return context


class RekapPasienTerapisView(PermissionRequiredViewMixin, TemplateView):
    """Summary of number of patients per therapist"""
    template_name = 'core/pembukuan/rekap_pasien_terapis.html'
    permission_required = 'pembukuan_view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get summary: count registrasi per terapis
        from django.db.models import Count
        terapis_list = Terapis.objects.annotate(
            jumlah_pasien=Count('registrasi')
        ).order_by('-jumlah_pasien')
        
        total_registrasi = Registrasi.objects.count()
        
        context['terapis_list'] = terapis_list
        context['total_terapis'] = terapis_list.count()
        context['total_registrasi'] = total_registrasi
        return context


class RekapTransportTerapisView(PermissionRequiredViewMixin, TemplateView):
    """Summary of transport money per therapist"""
    template_name = 'core/pembukuan/rekap_transport_terapis.html'
    permission_required = 'pembukuan_view'
    
    def get_context_data(self, **kwargs):
        import json
        context = super().get_context_data(**kwargs)
        
        # Get summary: total biaya_transport per terapis
        terapis_transport = []
        for terapis in Terapis.objects.all():
            registrasi = Registrasi.objects.filter(terapis=terapis)
            total_transport = registrasi.aggregate(Sum('biaya_transport'))['biaya_transport__sum'] or 0
            terapis_transport.append({
                'terapis': terapis,
                'total_transport': total_transport,
                'total_transport_formatted': format_rupiah(total_transport),
                'jumlah_registrasi': registrasi.count()
            })
        
        # Sort by total_transport descending
        terapis_transport.sort(key=lambda x: x['total_transport'], reverse=True)
        grand_total = sum(item['total_transport'] for item in terapis_transport)
        
        # Prepare chart data for distribution
        terapis_names = [item['terapis'].nama_terapis for item in terapis_transport]
        transport_amounts = [float(item['total_transport']) for item in terapis_transport]
        
        # Color palette for chart
        colors = [
            '#0d6efd', '#6c757d', '#198754', '#dc3545', '#ffc107',
            '#0dcaf0', '#6f42c1', '#d63384', '#fd7e14', '#20c997',
            '#e83e8c', '#17a2b8', '#007bff', '#28a745', '#dc3545',
        ]
        # Cycle colors if more terapis than colors
        chart_colors = [colors[i % len(colors)] for i in range(len(terapis_names))]
        
        context['terapis_transport'] = terapis_transport
        context['grand_total'] = grand_total
        context['grand_total_formatted'] = format_rupiah(grand_total)
        context['terapis_names_json'] = json.dumps(terapis_names)
        context['transport_amounts_json'] = json.dumps(transport_amounts)
        context['chart_colors_json'] = json.dumps(chart_colors)
        return context


class SaldoAkhirView(PermissionRequiredViewMixin, TemplateView):
    """Final balance report (Total Income - Total Expenses)"""
    template_name = 'core/pembukuan/saldo_akhir.html'
    permission_required = 'pembukuan_view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate totals
        total_pendapatan = Pemasukan.objects.aggregate(Sum('jumlah'))['jumlah__sum'] or 0
        total_pengeluaran = Pengeluaran.objects.aggregate(Sum('jumlah'))['jumlah__sum'] or 0
        saldo_akhir = total_pendapatan - total_pengeluaran
        
        context['total_pendapatan'] = total_pendapatan
        context['total_pendapatan_formatted'] = format_rupiah(total_pendapatan)
        context['total_pengeluaran'] = total_pengeluaran
        context['total_pengeluaran_formatted'] = format_rupiah(total_pengeluaran)
        context['saldo_akhir'] = saldo_akhir
        context['saldo_akhir_formatted'] = format_rupiah(saldo_akhir)
        context['status'] = 'surplus' if saldo_akhir > 0 else 'defisit' if saldo_akhir < 0 else 'seimbang'
        
        return context


# ============================================
# PENGATURAN (SETTINGS)
# ============================================

class AppSettingsView(PermissionRequiredViewMixin, View):
    """View for managing app settings (font size and logo)."""
    login_url = '/login/'
    template_name = 'core/pengaturan.html'
    permission_map = {
        'GET': 'settings_view',
        'POST': 'settings_edit',
    }

    def get(self, request):
        from .models import AppSettings
        from .forms import AppSettingsForm
        
        settings = AppSettings.get_settings()
        form = AppSettingsForm(instance=settings)
        
        context = {
            'form': form,
            'settings': settings,
        }
        return render(request, self.template_name, context)

    def post(self, request):
        from .models import AppSettings
        from .forms import AppSettingsForm
        
        settings = AppSettings.get_settings()
        form = AppSettingsForm(request.POST, request.FILES, instance=settings)
        
        if form.is_valid():
            settings = form.save(commit=False)
            settings.updated_by = request.user
            settings.save()
            messages.success(request, 'Pengaturan berhasil disimpan!')
            return redirect('app_settings')
        
        context = {
            'form': form,
            'settings': settings,
        }
        return render(request, self.template_name, context)


# ============================================
# USER MANAGEMENT
# ============================================

class RoleManagementAccessMixin(LoginRequiredMixin):
    login_url = '/login/'

    def dispatch(self, request, *args, **kwargs):
        sync_permission_catalog()
        if not can_manage_roles(request.user):
            raise PermissionDenied('Hanya superadmin yang dapat mengelola role dan privileges.')
        return super().dispatch(request, *args, **kwargs)


class RoleListView(RoleManagementAccessMixin, ListView):
    model = Role
    template_name = 'core/role_list.html'
    context_object_name = 'roles'

    def get_queryset(self):
        return Role.objects.all().annotate(
            total_permissions=Count('rolepermission__permission_id', distinct=True),
            total_users=Count('userrole__user_id', distinct=True),
        ).order_by('nama_role')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['role_blueprints'] = DEFAULT_ROLE_BLUEPRINTS
        context['bootstrap_mode'] = is_rbac_bootstrap_mode()
        return context


class RoleCreateView(RoleManagementAccessMixin, View):
    template_name = 'core/role_form.html'

    def get(self, request):
        form = RoleForm()
        context = {
            'form': form,
            'is_create': True,
            'permission_groups': form.permission_groups,
            'selected_permission_ids': form.get_selected_permission_ids(),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        form = RoleForm(request.POST)
        if form.is_valid():
            role = form.save()
            messages.success(request, f'Role {role.nama_role} berhasil dibuat.')
            return redirect('role_list')

        context = {
            'form': form,
            'is_create': True,
            'permission_groups': form.permission_groups,
            'selected_permission_ids': form.get_selected_permission_ids(),
        }
        return render(request, self.template_name, context)


class RoleEditView(RoleManagementAccessMixin, View):
    template_name = 'core/role_form.html'

    def get(self, request, pk):
        role = get_object_or_404(Role, pk=pk)
        form = RoleForm(instance=role)
        context = {
            'form': form,
            'is_create': False,
            'role_obj': role,
            'permission_groups': form.permission_groups,
            'selected_permission_ids': form.get_selected_permission_ids(),
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        role = get_object_or_404(Role, pk=pk)
        form = RoleForm(request.POST, instance=role)
        if form.is_valid():
            role = form.save()
            messages.success(request, f'Role {role.nama_role} berhasil diperbarui.')
            return redirect('role_list')

        context = {
            'form': form,
            'is_create': False,
            'role_obj': role,
            'permission_groups': form.permission_groups,
            'selected_permission_ids': form.get_selected_permission_ids(),
        }
        return render(request, self.template_name, context)


class RoleSeedDefaultsView(RoleManagementAccessMixin, View):
    def post(self, request):
        result = seed_default_roles()
        created = ', '.join(result['created']) or '-'
        updated = ', '.join(result['updated']) or '-'
        messages.success(request, f'Role rekomendasi berhasil disiapkan. Baru: {created}. Diperbarui: {updated}.')
        return redirect('role_list')

class UserListView(LoginRequiredMixin, ListView):
    """List all users."""
    model = User
    template_name = 'core/user_list.html'
    context_object_name = 'users'
    paginate_by = 25

    def get_queryset(self):
        return User.objects.select_related('cabang').order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        for user in context['users']:
            user.assigned_role_names = list(
                Role.objects.filter(userrole__user=user).values_list('nama_role', flat=True)
            )
        context['can_manage_roles'] = can_manage_roles(self.request.user)
        return context


class UserCreateView(LoginRequiredMixin, View):
    """Create new user."""
    login_url = '/login/'
    template_name = 'core/user_form.html'

    def get(self, request):
        form = UserCreateForm(current_user=request.user)
        context = {
            'form': form,
            'is_create': True,
            'can_manage_roles': can_manage_roles(request.user),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        form = UserCreateForm(request.POST, current_user=request.user)
        
        if form.is_valid():
            user = form.save()
            messages.success(request, f'User {user.username} berhasil dibuat!')
            return redirect('user_list')
        
        context = {
            'form': form,
            'is_create': True,
            'can_manage_roles': can_manage_roles(request.user),
        }
        return render(request, self.template_name, context)


class UserEditView(LoginRequiredMixin, View):
    """Edit existing user."""
    login_url = '/login/'
    template_name = 'core/user_form.html'

    def get(self, request, pk):
        user = User.objects.get(pk=pk)
        form = UserForm(instance=user, current_user=request.user)
        context = {
            'form': form,
            'is_create': False,
            'user_obj': user,
            'can_manage_roles': can_manage_roles(request.user),
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        user = User.objects.get(pk=pk)
        form = UserForm(request.POST, instance=user, current_user=request.user)
        
        if form.is_valid():
            user = form.save()
            messages.success(request, f'User {user.username} berhasil diupdate!')
            return redirect('user_list')
        
        context = {
            'form': form,
            'is_create': False,
            'user_obj': user,
            'can_manage_roles': can_manage_roles(request.user),
        }
        return render(request, self.template_name, context)


class UserToggleActiveView(LoginRequiredMixin, View):
    """Toggle user active status (activate/deactivate)."""
    login_url = '/login/'

    def post(self, request, pk):
        user = User.objects.get(pk=pk)
        user.is_active = not user.is_active
        user.save()
        
        status = 'diaktifkan' if user.is_active else 'dinonaktifkan'
        messages.success(request, f'User {user.username} berhasil {status}!')
        return redirect('user_list')


class ChangeOwnPasswordView(LoginRequiredMixin, FormView):
    """Allow the currently logged-in user to change their own password."""
    login_url = '/login/'
    template_name = 'core/change_password.html'
    form_class = UserPasswordChangeForm
    success_url = reverse_lazy('change_own_password')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        user = form.save()
        update_session_auth_hash(self.request, user)
        messages.success(self.request, 'Password berhasil diubah. Sesi login Anda tetap aktif.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Edit Password'
        return context

# ============================================
# NOTIFICATION GENERATION
# ============================================

class GenerateNotificationsView(LoginRequiredMixin, View):
    """Generate notifications on-demand."""
    login_url = '/login/'
    permission_required = 'notifikasi_generate'

    def post(self, request):
        if not request.user.has_permission(self.permission_required):
            return deny_permission_response(request, self.permission_required, 'Anda tidak punya akses untuk generate notifikasi.')
        from core.services.notification_service import generate_all_notifications
        from django.http import JsonResponse
        
        try:
            result = generate_all_notifications()
            messages.success(
                request,
                f"✓ Notifikasi berhasil dibuat! Total: {result['total_created']} notifikasi baru"
            )
            return JsonResponse({
                'success': True,
                'total_created': result['total_created'],
                'details': {k: v['message'] for k, v in result['details'].items()}
            })
        except Exception as e:
            messages.error(request, f'Error generating notifications: {str(e)}')
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

# ==========================================
# API Endpoint untuk Multi-Terapi
# ==========================================

def api_jenis_terapi_detail(request, pk):
    """API endpoint untuk get detail jenis terapi (untuk AJAX)"""
    try:
        terapi = JenisTerapi.objects.get(pk=pk, is_deleted=False)
        return JsonResponse({
            'id': terapi.id,
            'nama_terapi': terapi.nama_terapi,
            'harga': float(terapi.harga),
            'harga_formatted': f'Rp {int(terapi.harga):,}'.replace(',', '.')
        })
    except JenisTerapi.DoesNotExist:
        return JsonResponse({'error': 'Terapi tidak ditemukan'}, status=404)

# ==========================================
# CALENDAR VIEW
# ==========================================

class CalendarView(PermissionRequiredViewMixin, TemplateView):
    """Calendar view for appointment scheduling"""
    template_name = 'core/calendar.html'
    permission_required = 'calendar_view'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get all terapis for filter
        context['terapis_list'] = Terapis.objects.filter(is_active=True, is_deleted=False).order_by('nama_terapis')
        # Get all cabangs for filter
        context['cabang_list'] = Cabang.objects.all().order_by('nama_cabang')
        return context

@require_permission('calendar_view')
def calendar_events_api(request):
    """API endpoint for FullCalendar events (JSON)"""
    from datetime import datetime
    import json
    
    # Get query parameters
    start = request.GET.get('start')  # ISO format dari FullCalendar
    end = request.GET.get('end')
    terapis_id = request.GET.get('terapis_id')
    cabang_id = request.GET.get('cabang_id')
    status_filter = request.GET.get('status_filter')
    
    # Base queryset
    qs = Registrasi.objects.select_related('pasien', 'terapis', 'jenis_terapi', 'cabang').filter(is_deleted=False)
    
    # Filter by date range
    if start:
        try:
            start_date = datetime.fromisoformat(start.replace('Z', '+00:00')).date()
            qs = qs.filter(tanggal_kunjungan__gte=start_date)
        except:
            pass
    
    if end:
        try:
            end_date = datetime.fromisoformat(end.replace('Z', '+00:00')).date()
            qs = qs.filter(tanggal_kunjungan__lte=end_date)
        except:
            pass
    
    # Filter by terapis
    if terapis_id:
        qs = qs.filter(terapis_id=terapis_id)
    
    # Filter by cabang
    if cabang_id:
        qs = qs.filter(cabang_id=cabang_id)
    
    # Filter by status
    if status_filter:
        qs = qs.filter(status=status_filter)
    
    # Build events list for FullCalendar
    events = []
    for reg in qs:
        # Determine color based on status - VERY DARK for better contrast
        color_map = {
            'BOOKED': '#92400e',      # Very Dark Orange
            'CONFIRMED': '#1e40af',   # Very Dark Blue
            'COMPLETED': '#065f46',   # Very Dark Green
            'CANCELLED': '#991b1b',   # Very Dark Red
        }
        color = color_map.get(reg.status, '#1f2937')
        
        events.append({
            'id': reg.id,
            'title': f"{reg.pasien.nama_anak} - {reg.jenis_terapi.nama_terapi}",
            'start': reg.tanggal_kunjungan.isoformat(),
            'backgroundColor': color,
            'borderColor': color,
            'extendedProps': {
                'pasien': reg.pasien.nama_anak,
                'terapis': reg.terapis.nama_terapis if reg.terapis else 'Belum ditentukan',
                'terapis_id': reg.terapis.id if reg.terapis else None,
                'jenis_terapi': reg.jenis_terapi.nama_terapi,
                'status': reg.status,
                'cabang': reg.cabang.nama_cabang if reg.cabang else '-',
                'kode_registrasi': reg.kode_registrasi or '-',
            }
        })
    
    return JsonResponse(events, safe=False)

@require_permission('registrasi_reschedule')
@require_http_methods(["POST"])
def update_registrasi_date_api(request, registrasi_id):
    """
    API endpoint to update registrasi date via drag & drop
    """
    try:
        import json
        from datetime import datetime
        
        # Parse request body
        data = json.loads(request.body)
        new_date_str = data.get('tanggal_kunjungan')
        
        if not new_date_str:
            return JsonResponse({'success': False, 'message': 'Tanggal tidak valid'}, status=400)
        
        # Get registrasi
        registrasi = Registrasi.objects.get(id=registrasi_id, is_deleted=False)
        old_date = registrasi.tanggal_kunjungan
        today = timezone.localdate()

        if old_date < today:
            return JsonResponse({
                'success': False,
                'message': 'Jadwal yang sudah terlewat tidak bisa dipindahkan.'
            }, status=400)
        
        # Parse new date
        new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()

        if new_date <= today:
            return JsonResponse({
                'success': False,
                'message': 'Jadwal hanya bisa dipindah ke hari mendatang (H+).'
            }, status=400)
        
        # Update tanggal_kunjungan
        registrasi.tanggal_kunjungan = new_date
        registrasi.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Appointment berhasil dipindahkan dari {old_date.strftime("%d %b %Y")} ke {new_date.strftime("%d %b %Y")}',
            'old_date': old_date.isoformat(),
            'new_date': new_date.isoformat()
        })
        
    except Registrasi.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Registrasi tidak ditemukan'}, status=404)
    except ValueError as e:
        return JsonResponse({'success': False, 'message': f'Format tanggal invalid: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)
